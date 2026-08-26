import os
import io
import re
import json
import copy
import uuid
import asyncio
import datetime as dt
import logging
import imaplib
import smtplib
import email
import email.utils
from email.header import decode_header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

import httpx
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.utils import ImageReader
from pypdf import PdfReader, PdfWriter

try:
    # 구글 드라이브 자동 업로드용(설정 안 해두면 그냥 이 기능만 꺼짐)
    from google.oauth2 import service_account as _gdrive_service_account
    from googleapiclient.discovery import build as _gdrive_build
    from googleapiclient.http import MediaInMemoryUpload as _gdrive_media
except ImportError:
    _gdrive_service_account = None
    _gdrive_build = None
    _gdrive_media = None
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)
from anthropic import Anthropic

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
ALLOWED_USER_ID = int(os.environ["ALLOWED_USER_ID"])

GMAIL_ADDRESS = os.environ.get("GMAIL_ADDRESS")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD")
MAIL_CHECK_INTERVAL = int(os.environ.get("MAIL_CHECK_INTERVAL", "120"))

# 메일 보낼 때 "보낸사람"으로 표시할 주소. 지정 안 하면 GMAIL_ADDRESS 그대로 사용.
# (구글 계정에 "다른 이메일 주소로 보내기" 별칭으로 인증된 주소여야 정상 발송돼요)
MAIL_FROM_ADDRESS = os.environ.get("MAIL_FROM_ADDRESS", GMAIL_ADDRESS)

# 보낸사람은 그대로 GMAIL_ADDRESS로 두되(별도 별칭 인증 없이 바로 되는 방식), 담당자가
# "답장"을 누르면 유진님의 실제 업무 메일(네이버, 구글로 자동전달 중)로 가도록 지정하는 주소
MAIL_REPLY_TO = os.environ.get("MAIL_REPLY_TO", MAIL_FROM_ADDRESS)

MODEL_NAME = os.environ.get("MODEL_NAME", "claude-haiku-4-5-20251001")

# 가입증명서를 구글 드라이브(유진님 PC에 동기화된 폴더)에도 자동으로 올려주는 기능용 설정.
# 서비스 계정 방식이라 유진님이 매번 로그인/인증할 필요 없이, 처음에 딱 한 번 그 서비스
# 계정한테 드라이브 폴더 하나를 "공유"만 해두면 계속 자동으로 파일을 넣어줌.
GDRIVE_SERVICE_ACCOUNT_JSON = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON")
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID")

# 카카오 REST API 키 (카카오맵 대중교통/도보/자전거 길찾기 + 카카오모빌리티 자동차 길찾기 공용)
KAKAO_REST_API_KEY = os.environ.get("KAKAO_REST_API_KEY")

client = Anthropic(api_key=ANTHROPIC_API_KEY)

# 가입증명서 자동 생성용 자산(폰트/템플릿) 경로
# assets 폴더가 있으면 그 안에서, 없으면(루트에 바로 올린 경우) bot.py와 같은 위치에서 찾음
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 통합파일/증권번호/브랜드 설정처럼 "재배포해도 사라지면 안 되는" 데이터를 저장할 위치.
# Render에서 영구 디스크(Persistent Disk)를 만들어 PERSIST_DIR 환경변수로 그 마운트 경로를
# 지정해두면 그쪽에 저장하고, 설정 안 해두면(기본값) 지금처럼 코드 폴더에 저장함(재배포시 초기화).
PERSIST_DIR = os.environ.get("PERSIST_DIR", _BASE_DIR)
os.makedirs(PERSIST_DIR, exist_ok=True)

# 집/회사 등 저장된 위치를 담아두는 파일 (재시작해도 남아있도록 디스크에 저장)
PLACES_FILE = os.path.join(PERSIST_DIR, "places.json")


def _find_file_ci(directory: str, filename: str) -> str | None:
    """Render 서버(Linux)는 대소문자를 구분해서, GitHub 웹에서 파일 이름을 바꿀 때
    확장자가 의도치 않게 .PDF처럼 대문자로 바뀌어 있으면 .pdf로 찾다가 못 찾는 경우가 생김.
    그래서 폴더 안 파일 목록을 대소문자 구분 없이 한 번 더 확인함."""
    try:
        entries = os.listdir(directory)
    except OSError:
        return None
    target = filename.lower()
    for entry in entries:
        if entry.lower() == target:
            return os.path.join(directory, entry)
    return None


def _find_asset(filename: str) -> str:
    # GitHub 웹으로 파일을 올릴 때 이름 첫 글자가 의도치 않게 대문자로 바뀌는 경우가 있어서
    # (예: batang.ttc -> Batang.ttc), 정확한 이름으로 못 찾으면 대소문자 구분 없이 한 번 더 찾아봄
    for base in (os.path.join(_BASE_DIR, "assets"), _BASE_DIR):
        candidate = os.path.join(base, filename)
        if os.path.exists(candidate):
            return candidate
        ci_match = _find_file_ci(base, filename)
        if ci_match:
            return ci_match
    return os.path.join(_BASE_DIR, filename)


CERT_FONT_PATH = _find_asset("batang.ttc")
CERT_FONT_INDEX = 1  # batang.ttc 안에는 바탕/바탕체/궁서/궁서체 4종류가 들어있는데, 서식 PDF들이 실제로
# 쓰는 폰트가 '바탕체(BatangChe)'라서(pdfplumber로 서식 글자를 직접 확인함) 그 인덱스를 지정해서 씀
CERT_TEMPLATE_PATH = _find_asset("cert_template.pdf")  # 브랜드별 서식이 없을 때 쓰는 기본값(트레몰로 서식)

# 담당자에게 보내는 이메일 맨 아래에 붙는 명함/서명. 바뀌면 여기만 수정하면 됨.
_CERT_EMAIL_SIGNATURE = (
    "황유진 대표\n"
    "다온인슈 기업보험 전문컨설팅 / 대표이사\n\n"
    "우)06620 서울특별시 서초구 강남대로 375 서초현대타워 10층\n"
    "Tel 02-6953-8129 Mobile 010-9810-5503\n"
    "Fax 0303-0950-5503\n"
    "Email yujin.hwang@daonins.co.kr"
)

# 담당자에게 가입증명서를 이메일로 보낼 때 본문 문구. {store_name} 자리에 매장명이 자동으로 들어감.
# 문구를 바꾸고 싶으면 여기만 수정하면 됨.
_CERT_EMAIL_BODY_TEMPLATE = (
    "안녕하십니까.\n\n"
    "요청해주신 '{store_name}' 신규 매장 가입증명서를 첨부와 같이 보내드립니다.\n\n"
    "확인 부탁드리며, 문의사항 있으시면 언제든 연락 주시기 바랍니다.\n\n"
    "감사합니다.\n\n"
    "--\n" + _CERT_EMAIL_SIGNATURE
)
CERT_PAGE_W, CERT_PAGE_H = 595.2, 841.92
CERT_FONT_SCALE = 4  # 텍스트를 이미지로 그릴 때 선명하게 보이도록 확대 비율

# 브랜드별 가입증명서 서식에서 5개 항목(보험기간/점포명/주소/보상한도액/보험료) 줄의 세로 위치(top, pt)와
# '보험료' 값 칸의 오른쪽 한계(premium_x1 - 바로 뒤 '원' 글자를 안 지우면서 값을 덮을 수 있는 최대 x좌표).
# 같은 디자인이라도 브랜드마다 '계약자'/'주소' 줄 길이가 달라 아래 항목들이 조금씩 밀려 있고, '원' 글자
# 시작 위치도 살짝 달라서, 실제 발급된 예시 증명서에서 좌표를 직접 재서 브랜드별로 따로 저장해둠
# (pdfplumber로 측정). premium_x1을 너무 좁게 잡으면 새로 쓰는 금액이 기존 서식 금액보다 짧을 때
# 예전 숫자 일부가 안 가려지고 남아 보이는 문제가 생김.
DEFAULT_CERT_ROWS = {"period": 201.4, "store_name": 295.3, "address": 313.3, "limit": 367.6, "premium": 439.8, "premium_x1": 160}
CERT_ROWS_BY_BRAND = {
    "트레몰로": {"period": 201.4, "store_name": 295.3, "address": 313.3, "limit": 367.6, "premium": 439.8, "premium_x1": 162.7},
    "월메이드": {"period": 183.4, "store_name": 277.1, "address": 300.6, "limit": 360.1, "premium": 432.4, "premium_x1": 168.7},
    "올리비아로렌&오뷔엘알&주얼리": {"period": 183.4, "store_name": 277.1, "address": 295.3, "limit": 349.6, "premium": 421.8, "premium_x1": 168.7},
}

# 브랜드별 정산 통합파일을 서버에 계속 보관/갱신하는 폴더
MASTERS_DIR = os.path.join(PERSIST_DIR, "masters")
# 브랜드별 증권번호를 저장해두는 파일 (파일명에서 못 찾을 때 사용)
POLICY_NUMBERS_FILE = os.path.join(PERSIST_DIR, "policy_numbers.json")
# 정산양식을 보내는 담당자 이름 -> 이메일 주소. 신규매장 가입증명서를 그 담당자에게
# 자동으로 이메일로 보내줄 때 씀(엑셀 '접수자' 칸의 이름으로 찾음)
CONTACTS_FILE = os.path.join(PERSIST_DIR, "contacts.json")
# 담당자에게 이메일로 보내기 전에 유진님 확인을 거치는 가입증명서들을 보관하는 곳.
# "보내기"를 누르기 전까지는 실제로 발송되지 않고 여기에 대기함(재배포/재시작해도
# 안 없어지도록 디스크에 저장).
PENDING_CERTS_DIR = os.path.join(PERSIST_DIR, "pending_certs")
PENDING_CERTS_INDEX_FILE = os.path.join(PERSIST_DIR, "pending_certs.json")
# 브랜드별로 파일을 보내드릴 때 쓸 표시용 파일명(엑셀 A1 셀 이름과 다르게 쓰고 싶을 때)
BRAND_NAMES_FILE = os.path.join(PERSIST_DIR, "brand_names.json")
# 담당자마다 A1 셀에 브랜드명을 다르게 적어 보내는 경우, 같은 브랜드로 취급하도록 이름을 통일시켜주는 매핑
BRAND_ALIASES_FILE = os.path.join(PERSIST_DIR, "brand_aliases.json")
# 브랜드별 가입증명서 '원본 서식' PDF 폴더. 실제로 발급된 예시 증명서를 그대로 서식으로 써서,
# 증권번호/계약자/보험종목 등은 그 서식에 이미 올바르게 적혀 있는 값을 그대로 두고,
# 매장마다 달라지는 5개 항목(보험기간/점포명/주소/보상한도액/보험료)만 덮어써서 새로 만듦.
# assets/cert_templates 에 미리 등록해둔 브랜드(트레몰로/월메이드/올리비아로렌&오뷔엘알&주얼리)는
# 코드와 함께 배포되어 재배포해도 사라지지 않음. 새 브랜드는 /setcerttemplate 로 직접 추가 가능(이땐
# 서버 디스크에 저장되어 재배포 시 다시 등록해야 함). /setcerttemplate로 등록한 서식은
# PERSIST_DIR(영구 디스크)에 저장해서 재배포해도 남아있게 함.
CERT_TEMPLATES_DIR = os.path.join(PERSIST_DIR, "cert_templates")
CERT_BUNDLED_TEMPLATES_DIR = os.path.join(_BASE_DIR, "assets", "cert_templates")

SYSTEM_PROMPT = (
    "당신은 사용자의 개인 AI 비서입니다. 한국어로 친절하고 간결하게 답변하세요. "
    "불필요하게 길게 설명하지 말고, 핵심 위주로 답하세요. "
    "날씨, 최신 뉴스, 맛집, 가격 등 최신 정보가 필요한 질문은 웹 검색 도구를 적극 활용하세요."
)

CHAT_TOOLS = [
    {"type": "web_search_20250305", "name": "web_search", "max_uses": 3}
]

conversations: dict[int, list[dict]] = {}
MAX_HISTORY = 20

# 메일 감시용: 마지막으로 확인한 IMAP UID (재시작하면 초기화됨)
last_uid_seen: int | None = None


def is_allowed(update: Update) -> bool:
    user = update.effective_user
    return user is not None and user.id == ALLOWED_USER_ID


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        await update.message.reply_text("죄송해요, 이 봇은 개인용이라 사용할 수 없어요.")
        return
    await update.message.reply_text(
        "안녕하세요! AI 비서예요. 👋\n\n"
        "그냥 편하게 메시지를 보내면 대화할 수 있어요.\n\n"
        "사용 가능한 명령어:\n"
        "/remind <분> <내용> - 알림 예약 (예: /remind 30 회의 참석)\n"
        "/mail <받는사람이메일> - 메일 작성해서 보내기 (파일 첨부 가능)\n"
        "/route - 두 지점 사이 길찾기 (자동차/대중교통/도보/자전거 비교)\n"
        "/sethome - 집 위치 등록\n"
        "/setwork - 회사 위치 등록\n"
        "/towork - 집→회사 길찾기 (등록 필요)\n"
        "/tohome - 회사→집 길찾기 (등록 필요)\n"
        "/reset - 지금까지의 대화 기억 지우기\n"
        "/setpolicy <브랜드명> <증권번호> - 브랜드별 증권번호 등록 (가입증명서에 사용)\n"
        "/brands - 등록된 브랜드(통합파일) 목록 확인\n"
        "/resetbrand <브랜드명> - 해당 브랜드 통합파일 삭제하고 처음부터 다시 등록\n"
        "/setbrandname <브랜드명> <표시할 이름> - 통합파일을 보내드릴 때 쓸 파일명 변경\n"
        "/sendmaster <브랜드명> - 지금 저장된 통합파일을 새로 올리지 않고 다시 받아보기\n"
        "/setbrandalias <다르게 인식된 이름> <진짜 브랜드명> - 담당자마다 다르게 적는 브랜드명을 하나로 통일\n"
        "/setcerttemplate <브랜드명> - (예시 PDF에 답장하며 사용) 그 브랜드 전용 가입증명서 서식으로 등록\n"
        "/setcontact <담당자 이름> <이메일> - 담당자 이메일 등록 (신규매장 가입증명서 발송용)\n"
        "/contacts - 등록된 담당자 연락처 목록 보기\n"
        "/pending - 대기 중인(아직 안 보낸) 가입증명서 목록 보기\n"
        "/testmail <받는이메일> - 샘플 가입증명서로 메일 발송 테스트 (받는이메일 생략하면 내 메일로)\n\n"
        "길찾기는 명령어 없이 그냥 '강남역까지 얼마나 걸려?', '홍대에서 여의도까지 어떻게 가?'처럼 물어보셔도 알아들어요.\n"
        "가입증명서도 'OO점 가입증명서 찾아줘'처럼 편하게 말하면 등록된 매장 중에서 찾아서 다시 보내드려요.\n\n"
        "새 이메일이 오면 자동으로 요약해서 알려드려요. 📬\n"
        "메일에 정산양식 엑셀이 첨부되어 있으면, 다운로드 안 하셔도 자동으로 확인해서 신규/폐점 매장을 반영하고 가입증명서와 갱신된 통합파일을 보내드려요.\n"
        "날씨, 최신 뉴스, 맛집 등도 그냥 물어보시면 웹 검색해서 답해드려요.\n"
        "📎(첨부) 버튼으로 '위치'를 공유해주시면, 그 위치 기준으로 근처 맛집도 찾아드려요.\n\n"
        "정산양식 엑셀 파일을 보내주시면, 브랜드별 통합파일과 비교해서 새로 추가된 매장을 찾아 가입증명서 PDF를 자동으로 만들어드리고, 갱신된 통합파일도 함께 보내드려요. 📄"
    )


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    conversations.pop(update.effective_chat.id, None)
    await update.message.reply_text("대화 기록을 초기화했어요.")


async def remind(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return

    try:
        minutes = float(context.args[0])
        text = " ".join(context.args[1:]).strip()
        if not text:
            raise ValueError
    except (IndexError, ValueError):
        await update.message.reply_text(
            "사용법: /remind <분> <내용>\n예: /remind 30 회의 참석"
        )
        return

    chat_id = update.effective_chat.id
    context.job_queue.run_once(
        send_reminder, when=minutes * 60, chat_id=chat_id, data=text
    )
    await update.message.reply_text(f"⏰ {minutes}분 후에 알려드릴게요: {text}")


async def send_reminder(context: ContextTypes.DEFAULT_TYPE) -> None:
    job = context.job
    await context.bot.send_message(chat_id=job.chat_id, text=f"⏰ 알림: {job.data}")


async def _reverse_geocode(lat: float, lon: float) -> str:
    url = "https://nominatim.openstreetmap.org/reverse"
    params = {"lat": lat, "lon": lon, "format": "json", "accept-language": "ko"}
    headers = {"User-Agent": "yujin-ai-bot/1.0"}
    async with httpx.AsyncClient(timeout=10) as http_client:
        resp = await http_client.get(url, params=params, headers=headers)
        resp.raise_for_status()
        data = resp.json()
    return data.get("display_name") or f"위도 {lat}, 경도 {lon}"


def _load_places() -> dict:
    try:
        with open(PLACES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_places(places: dict) -> None:
    with open(PLACES_FILE, "w", encoding="utf-8") as f:
        json.dump(places, f, ensure_ascii=False, indent=2)


async def _kakao_address_search(query: str) -> dict | None:
    """정확한 주소 문자열을 좌표로 변환 (지번/도로명 주소 전용, 상세 동/호수는 제외 권장)"""
    if not KAKAO_REST_API_KEY:
        return None
    url = "https://dapi.kakao.com/v2/local/search/address.json"
    headers = {"Authorization": f"KakaoAK {KAKAO_REST_API_KEY}"}
    params = {"query": query}
    async with httpx.AsyncClient(timeout=10) as http_client:
        resp = await http_client.get(url, headers=headers, params=params)
        resp.raise_for_status()
        data = resp.json()
    docs = data.get("documents") or []
    if not docs:
        return None
    d = docs[0]
    return {
        "name": d.get("address_name"),
        "address": d.get("address_name"),
        "x": d["x"],
        "y": d["y"],
    }


async def _kakao_geocode_keyword(query: str) -> dict | None:
    """주소나 장소명을 좌표로 변환 (카카오맵 키워드 검색)"""
    if not KAKAO_REST_API_KEY:
        return None
    url = "https://dapi.kakao.com/v2/local/search/keyword.json"
    headers = {"Authorization": f"KakaoAK {KAKAO_REST_API_KEY}"}
    params = {"query": query, "size": 1}
    async with httpx.AsyncClient(timeout=10) as http_client:
        resp = await http_client.get(url, headers=headers, params=params)
        resp.raise_for_status()
        data = resp.json()
    docs = data.get("documents") or []
    if not docs:
        return None
    d = docs[0]
    return {
        "name": d.get("place_name") or d.get("road_address_name") or d.get("address_name"),
        "address": d.get("road_address_name") or d.get("address_name"),
        "x": d["x"],
        "y": d["y"],
    }


async def _extract_place_name_llm(text: str) -> str | None:
    """문장에서 장소명/주소만 뽑아냄 ('현재 위치는 강남역이야' -> '강남역')"""
    system = (
        "사용자 문장에서 언급된 장소명이나 주소만 정확히 추출해서 그 텍스트만 답하세요. "
        "다른 설명, 문장부호, 따옴표 없이 장소명만 출력하세요. "
        "장소를 찾을 수 없으면 정확히 NONE 이라고만 답하세요."
    )
    try:
        response = client.messages.create(
            model=MODEL_NAME,
            max_tokens=30,
            system=system,
            messages=[{"role": "user", "content": text}],
        )
        raw = "".join(b.text for b in response.content if b.type == "text").strip()
        if not raw or raw.upper() == "NONE":
            return None
        return raw
    except Exception:
        logger.exception("장소명 추출 중 오류")
        return None


async def _resolve_place(text: str, places: dict) -> dict | None:
    """사용자가 입력한 텍스트를 위치 정보로 변환. '집'/'회사'는 저장된 위치 사용"""
    t = text.strip()
    if t in ("집", "우리집", "집으로", "자택"):
        return places.get("home")
    if t in ("회사", "직장", "사무실", "회사로"):
        return places.get("work")
    try:
        # 정확한 주소(도로명/지번)는 주소 검색 API가 훨씬 정확함. 안 되면 장소명 검색으로 재시도
        place = await _kakao_address_search(t)
        if place:
            return place
        place = await _kakao_geocode_keyword(t)
        if place:
            return place
        # 문장 형태로 입력한 경우("현재 위치는 강남역이야") 장소명만 뽑아서 재시도
        extracted = await _extract_place_name_llm(t)
        if extracted and extracted != t:
            place = await _kakao_address_search(extracted)
            if place:
                return place
            return await _kakao_geocode_keyword(extracted)
        return None
    except Exception:
        logger.exception("장소 검색 중 오류")
        return None


async def _kakao_driving_route(sx: str, sy: str, ex: str, ey: str) -> dict | None:
    if not KAKAO_REST_API_KEY:
        return None
    url = "https://apis-navi.kakaomobility.com/v1/directions"
    headers = {"Authorization": f"KakaoAK {KAKAO_REST_API_KEY}"}
    params = {
        "origin": f"{sx},{sy}",
        "destination": f"{ex},{ey}",
        "priority": "RECOMMEND",
    }
    try:
        async with httpx.AsyncClient(timeout=10) as http_client:
            resp = await http_client.get(url, headers=headers, params=params)
            resp.raise_for_status()
            data = resp.json()
        routes = data.get("routes") or []
        if not routes or routes[0].get("result_code") != 0:
            return None
        return routes[0].get("summary")
    except Exception:
        logger.exception("자동차 길찾기 조회 중 오류")
        return None


async def _kakao_transit_route(sx: str, sy: str, ex: str, ey: str) -> dict | None:
    if not KAKAO_REST_API_KEY:
        return None
    url = "https://dapi.kakao.com/v2/routing/publictraffic"
    headers = {"Authorization": f"KakaoAK {KAKAO_REST_API_KEY}"}
    params = {"start_x": sx, "start_y": sy, "end_x": ex, "end_y": ey}
    try:
        async with httpx.AsyncClient(timeout=10) as http_client:
            resp = await http_client.get(url, headers=headers, params=params)
            resp.raise_for_status()
            return resp.json()
    except Exception:
        logger.exception("대중교통 길찾기 조회 중 오류")
        return None


async def _kakao_walk_or_bicycle_route(sx: str, sy: str, ex: str, ey: str, mode: str) -> dict | None:
    if not KAKAO_REST_API_KEY:
        return None
    path = "walk" if mode == "walk" else "bicycle"
    url = f"https://dapi.kakao.com/v2/routing/{path}"
    headers = {"Authorization": f"KakaoAK {KAKAO_REST_API_KEY}"}
    params = {"start_x": sx, "start_y": sy, "end_x": ex, "end_y": ey}
    try:
        async with httpx.AsyncClient(timeout=10) as http_client:
            resp = await http_client.get(url, headers=headers, params=params)
            resp.raise_for_status()
            return resp.json()
    except Exception:
        logger.exception("도보/자전거 길찾기 조회 중 오류")
        return None


def _format_driving(summary: dict | None) -> str:
    if not summary:
        return "🚗 자동차: 정보 없음 (승인 대기 중이거나 경로를 찾지 못했어요)"
    minutes = round(summary.get("duration", 0) / 60)
    km = round(summary.get("distance", 0) / 1000, 1)
    text = f"🚗 자동차: 약 {minutes}분 ({km}km)"
    toll = (summary.get("fare") or {}).get("toll")
    if toll:
        text += f", 통행료 {toll}원"
    return text


def _format_transit(data: dict | None) -> str:
    if not data or data.get("status") != "OK" or not data.get("routes"):
        return "🚌 대중교통: 정보 없음"
    route = data["routes"][0]
    props = route["properties"]
    minutes = round(props.get("totalTime", 0) / 60)
    line = f"🚌 대중교통: 약 {minutes}분"
    if props.get("transfers") is not None:
        line += f", 환승 {props['transfers']}회"
    fare = (props.get("fare") or {}).get("value")
    if fare:
        line += f", 요금 {fare}원"

    vehicle_names = []
    for step in route.get("steps", []):
        sp = step.get("properties", {})
        if sp.get("type") in ("BUS", "SUBWAY"):
            kind = "버스" if sp["type"] == "BUS" else "지하철"
            for v in sp.get("vehicles", []):
                label = f"{kind} {v.get('name', '')}".strip()
                if label not in vehicle_names:
                    vehicle_names.append(label)
    if vehicle_names:
        line += "\n   이용 노선: " + ", ".join(vehicle_names)
    return line


def _format_walk_or_bicycle(data: dict | None, label: str, emoji: str) -> str:
    if not data or data.get("status") != "OK":
        return f"{emoji} {label}: 정보 없음"
    props = data["route"]["properties"]
    minutes = round(props.get("totalTime", 0) / 60)
    km = round(props.get("totalDistance", 0) / 1000, 1)
    return f"{emoji} {label}: 약 {minutes}분 ({km}km)"


async def _reply_route_comparison(update: Update, start: dict, end: dict) -> None:
    start_label = start.get("name") or start.get("address") or "출발지"
    end_label = end.get("name") or end.get("address") or "도착지"

    if not KAKAO_REST_API_KEY:
        await update.message.reply_text("길찾기 기능이 아직 설정되지 않았어요.")
        return

    await update.message.reply_text(f"🔎 '{start_label}' → '{end_label}' 경로를 찾고 있어요...")

    sx, sy, ex, ey = start["x"], start["y"], end["x"], end["y"]

    driving, transit, walk, bicycle = await asyncio.gather(
        _kakao_driving_route(sx, sy, ex, ey),
        _kakao_transit_route(sx, sy, ex, ey),
        _kakao_walk_or_bicycle_route(sx, sy, ex, ey, "walk"),
        _kakao_walk_or_bicycle_route(sx, sy, ex, ey, "bicycle"),
    )

    lines = [
        f"📍 {start_label} → {end_label}",
        "",
        _format_driving(driving),
        _format_transit(transit),
        _format_walk_or_bicycle(walk, "도보", "🚶"),
        _format_walk_or_bicycle(bicycle, "자전거", "🚴"),
    ]
    await update.message.reply_text("\n".join(lines))


async def _advance_route_query(update: Update, context: ContextTypes.DEFAULT_TYPE, place: dict) -> None:
    """진행 중인 route_query 상태에 place를 반영하고 다음 단계로 진행 (텍스트 입력/위치 공유 공통 처리)"""
    route_query = context.user_data["route_query"]
    state = route_query["state"]
    label = place.get("name") or place.get("address") or "위치"

    if state == "await_current_location":
        dest_text = route_query.pop("pending_destination_text", None)
        if dest_text:
            places = _load_places()
            dest_place = await _resolve_place(dest_text, places)
            context.user_data.pop("route_query", None)
            if not dest_place:
                await update.message.reply_text(
                    f"현재 위치는 확인했는데, '{dest_text}'는 찾지 못했어요. 목적지를 다시 알려주세요."
                )
                context.user_data["route_query"] = {"state": "to", "from": place}
                return
            await _reply_route_comparison(update, place, dest_place)
        else:
            route_query["from"] = place
            route_query["state"] = "to"
            await update.message.reply_text(f"현재 위치: {label}\n\n도착지는 어디인가요?")
        return

    if state == "from":
        route_query["from"] = place
        route_query["state"] = "to"
        await update.message.reply_text(f"출발지: {label}\n\n도착지는 어디인가요?")
        return

    # state == "to"
    context.user_data.pop("route_query", None)
    await _reply_route_comparison(update, route_query["from"], place)


ROUTE_KEYWORDS = (
    "길찾", "가는길", "가는법", "가는방법", "어떻게가", "어떻게가나",
    "얼마나걸려", "얼마나걸리", "출근길", "퇴근길", "이동시간", "차로가", "대중교통으로",
    "버스로가", "도보로가", "걸어서가", "경로",
)


def _has_route_keyword(text: str) -> bool:
    normalized = text.replace(" ", "")
    return any(kw in normalized for kw in ROUTE_KEYWORDS)


async def _extract_route_intent(text: str) -> dict | None:
    """자유롭게 쓴 문장에서 길찾기 요청인지, 출발지/목적지가 무엇인지 뽑아냄"""
    system = (
        "사용자 문장이 길찾기(이동 경로) 요청인지 판단하고, 아래 JSON 형식으로만 답하세요. "
        "설명이나 다른 텍스트는 절대 포함하지 마세요.\n"
        '{"is_route": true 또는 false, '
        '"origin": "current"(현재 위치를 말하거나 출발지가 명시되지 않은 경우) 또는 "home" 또는 "work" 또는 구체적 장소명 문자열 또는 null, '
        '"destination": 구체적 장소명 문자열 또는 null(불명확한 경우)}\n'
        "길찾기 요청이 아니면 is_route를 false로 하세요."
    )
    try:
        response = client.messages.create(
            model=MODEL_NAME,
            max_tokens=200,
            system=system,
            messages=[{"role": "user", "content": text}],
        )
        raw = "".join(b.text for b in response.content if b.type == "text").strip()
        raw = raw.strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].strip()
        data = json.loads(raw)
        if not isinstance(data, dict):
            return None
        return data
    except Exception:
        logger.exception("길찾기 의도 분석 중 오류")
        return None


async def _handle_natural_route_request(update: Update, context: ContextTypes.DEFAULT_TYPE, user_text: str) -> bool:
    """자연어 길찾기 요청을 감지해서 처리. 처리했으면 True 반환"""
    if not KAKAO_REST_API_KEY:
        return False
    if not _has_route_keyword(user_text):
        return False

    intent = await _extract_route_intent(user_text)
    if not intent or not intent.get("is_route"):
        return False

    places = _load_places()
    origin = intent.get("origin")
    destination_text = intent.get("destination")

    # 출발지가 "현재 위치"를 가리키는 경우: 최근 공유된 위치가 있으면 재사용, 없으면 위치 공유 요청
    if origin in (None, "current"):
        recent = context.user_data.get("location")
        if isinstance(recent, dict) and recent.get("x") and recent.get("y"):
            from_place = {
                "name": recent.get("address"),
                "address": recent.get("address"),
                "x": recent["x"],
                "y": recent["y"],
            }
        else:
            context.user_data["route_query"] = {
                "state": "await_current_location",
                "pending_destination_text": destination_text,
            }
            extra = f" (목적지: {destination_text})" if destination_text else ""
            await update.message.reply_text(
                f"현재 위치를 확인할게요.{extra}\n📎(첨부) 버튼으로 위치를 공유해주세요."
            )
            return True
    else:
        from_place = await _resolve_place(origin, places)
        if not from_place:
            context.user_data["route_query"] = {"state": "from"}
            await update.message.reply_text(
                f"'{origin}' 위치를 찾지 못했어요. 출발지를 다시 알려주세요."
            )
            return True

    if not destination_text:
        context.user_data["route_query"] = {"state": "to", "from": from_place}
        await update.message.reply_text("도착지는 어디인가요?")
        return True

    to_place = await _resolve_place(destination_text, places)
    if not to_place:
        context.user_data["route_query"] = {"state": "to", "from": from_place}
        await update.message.reply_text(
            f"'{destination_text}'를 찾지 못했어요. 도착지를 다시 알려주세요."
        )
        return True

    await _reply_route_comparison(update, from_place, to_place)
    return True


async def route_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if not KAKAO_REST_API_KEY:
        await update.message.reply_text("길찾기 기능이 아직 설정되지 않았어요.")
        return
    context.user_data["route_query"] = {"state": "from"}
    await update.message.reply_text(
        "출발지를 알려주세요.\n"
        "주소나 장소명을 입력하거나, 📎(첨부)로 위치를 공유하거나, '집' 또는 '회사'라고 입력해주세요."
    )


async def set_home(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    context.user_data["saving_place"] = "home"
    await update.message.reply_text("🏠 집 위치를 설정할게요. 위치를 공유하거나 주소/건물명을 입력해주세요.")


async def set_work(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    context.user_data["saving_place"] = "work"
    await update.message.reply_text("🏢 회사 위치를 설정할게요. 위치를 공유하거나 주소/건물명을 입력해주세요.")


async def commute_to_work(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    places = _load_places()
    if not places.get("home") or not places.get("work"):
        await update.message.reply_text(
            "먼저 /sethome 과 /setwork 으로 집과 회사 위치를 등록해주세요."
        )
        return
    await _reply_route_comparison(update, places["home"], places["work"])


async def commute_to_home(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    places = _load_places()
    if not places.get("home") or not places.get("work"):
        await update.message.reply_text(
            "먼저 /sethome 과 /setwork 으로 집과 회사 위치를 등록해주세요."
        )
        return
    await _reply_route_comparison(update, places["work"], places["home"])


async def handle_location(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return

    loc = update.message.location
    try:
        address = await _reverse_geocode(loc.latitude, loc.longitude)
    except Exception:
        logger.exception("위치 확인 중 오류")
        address = f"위도 {loc.latitude}, 경도 {loc.longitude}"

    place = {"name": address, "address": address, "x": str(loc.longitude), "y": str(loc.latitude)}

    # 집/회사 위치 저장 중이면 그쪽으로 처리
    saving = context.user_data.get("saving_place")
    if saving:
        places = _load_places()
        places[saving] = place
        _save_places(places)
        context.user_data.pop("saving_place", None)
        label = "집" if saving == "home" else "회사"
        await update.message.reply_text(f"✅ {label} 위치를 저장했어요: {address}")
        return

    # 길찾기 진행 중이면 그쪽으로 처리
    if context.user_data.get("route_query"):
        await _advance_route_query(update, context, place)
        return

    context.user_data["location"] = {"address": address, "x": str(loc.longitude), "y": str(loc.latitude)}
    await update.message.reply_text(
        f"📍 위치를 받았어요: {address}\n"
        "이제 '근처 맛집 추천해줘'처럼 물어보시면 이 위치를 기준으로 찾아드릴게요."
    )


# ===================== 가입증명서 자동 생성 =====================
# 정산양식 엑셀(신규매장/폐점매장 시트)을 브랜드별 통합파일과 비교해 새로 추가된 매장을
# 찾아, 예시 PDF와 같은 양식의 가입증명서를 자동으로 만들어주는 기능. 원본 서식(배경/직인
# 이미지)은 그대로 두고 매장별로 달라지는 값(점포명/주소/보험기간/보험료 등)만 흰색으로
# 덮은 뒤 새로 그려넣는 방식.

# 브랜드에 등록된 증권번호가 없을 때 예시 양식과 동일하게 쓸 기본 증권번호
DEFAULT_POLICY_NO = "82509565736000"


def _cert_bundled_template_candidates(brand: str) -> list:
    """코드와 함께 배포되어 재배포해도 안 사라지는 브랜드별 서식을 찾을 수 있는 후보 (디렉터리, 파일명) 목록.
    GitHub에 assets 폴더를 따로 안 만들고 bot.py 등과 같은 위치(루트)에 그냥 올리는 경우가 많아서,
    assets/cert_templates 안과 루트(cert_template_브랜드명.pdf) 둘 다 확인함."""
    return [
        (CERT_BUNDLED_TEMPLATES_DIR, f"{brand}.pdf"),
        (_BASE_DIR, f"cert_template_{brand}.pdf"),
    ]


def _resolve_cert_bundled_template(brand: str) -> str | None:
    for directory, filename in _cert_bundled_template_candidates(brand):
        exact = os.path.join(directory, filename)
        if os.path.exists(exact):
            return exact
        # GitHub 웹에서 파일명을 바꿀 때 확장자가 .PDF로 바뀌어 있는 경우가 있어 대소문자 구분 없이도 확인
        found = _find_file_ci(directory, filename)
        if found:
            return found
    return None


def _cert_template_path_for_brand(brand: str) -> str:
    """브랜드별 가입증명서 원본 서식 경로. 순서: 사용자가 /setcerttemplate로 직접 등록한 것(서버 디스크,
    재배포시 사라짐) -> 코드와 함께 배포되는 것(재배포에도 안전) -> 아무것도 없으면 기본(트레몰로) 서식."""
    custom_path = os.path.join(CERT_TEMPLATES_DIR, f"{brand}.pdf")
    if os.path.exists(custom_path):
        return custom_path
    bundled = _resolve_cert_bundled_template(brand)
    if bundled:
        return bundled
    return CERT_TEMPLATE_PATH


def _cert_rows_for_brand(brand: str) -> dict:
    return CERT_ROWS_BY_BRAND.get(brand, DEFAULT_CERT_ROWS)


def _has_cert_template(brand: str) -> bool:
    """이 브랜드 전용 가입증명서 서식이 등록되어 있는지. 없으면 트레몰로 서식을 임시로 써서
    계약자/증권번호 등이 실제와 다르게 나올 수 있어 사용자에게 알려줘야 함."""
    if os.path.exists(os.path.join(CERT_TEMPLATES_DIR, f"{brand}.pdf")):
        return True
    return _resolve_cert_bundled_template(brand) is not None


# 가입증명서용 폰트(NotoSerifKR)에 글자가 비어있는(빈 도형) 특수문자들을 폰트가
# 지원하는 형태로 미리 바꿔줌. 안 하면 그 글자만 화면/PDF에서 통째로 빈 칸으로 사라짐.
# (예: '계약자' 줄에 '㈜세정'처럼 회사명 앞에 동그라미 표시가 들어가는 경우)
FONT_CHAR_FALLBACKS = {
    "㈜": "(주)",
}


def _sanitize_for_font(text: str) -> str:
    for bad, good in FONT_CHAR_FALLBACKS.items():
        text = text.replace(bad, good)
    return text


def _render_text_image(text: str, font_size_pt: float):
    text = _sanitize_for_font(text)
    px_size = int(font_size_pt * CERT_FONT_SCALE)
    font = ImageFont.truetype(CERT_FONT_PATH, px_size, index=CERT_FONT_INDEX)
    ascent, descent = font.getmetrics()
    bbox = font.getbbox(text)
    text_w = bbox[2] - bbox[0]
    img_h = ascent + descent
    img = Image.new("RGBA", (max(text_w, 1) + 4, img_h), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    draw.text((-bbox[0] + 2, 0), text, font=font, fill=(0, 0, 0, 255))
    return img, img.width / CERT_FONT_SCALE, img.height / CERT_FONT_SCALE, descent / CERT_FONT_SCALE


def _build_certificate_pdf(data: dict, brand: str) -> bytes:
    """브랜드별 실제 발급 예시 증명서를 원본 서식으로 그대로 쓰고, 매장마다 달라지는 5개 항목만
    흰색으로 덮은 뒤 새로 그려넣음. 증권번호/계약자/보험종목 등 나머지는 서식에 이미 올바르게
    적혀 있으므로 건드리지 않음(브랜드마다 회사명 문구가 달라도 신경 쓸 필요 없음).
    data keys: start_date, end_date, store_name, address, stock_amt(int), facility_amt(int), premium(int)"""
    rows = _cert_rows_for_brand(brand)
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=(CERT_PAGE_W, CERT_PAGE_H))

    def y_of(bottom):
        return CERT_PAGE_H - bottom

    def cover(x0, top, bottom, x1=545):
        c.setFillColorRGB(1, 1, 1)
        c.rect(x0 - 2, CERT_PAGE_H - bottom - 2, (x1 - x0) + 4, (bottom - top) + 4, fill=1, stroke=0)

    def draw_text(x0, bottom, s, size=12):
        # 이미지 자체가 이미 '글꼴 상 맨 아래(descent)'까지 포함해서 그려지므로, 이미지의 맨 아래를
        # 그 줄의 바텀 위치(baseline_y)에 바로 맞추면 됨. 예전에는 여기서 descent_pt를 한 번 더 빼서
        # 삽입한 글자들이 서식 원문 글자보다 항상 살짝(약 2pt) 아래로 처지는 문제가 있었음(바탕체로
        # 바꾸면서 서식과 나란히 놓고 보니 확실히 티가 남 — 사용자 확인 사례).
        img, w_pt, h_pt, descent_pt = _render_text_image(s, size)
        baseline_y = y_of(bottom)
        c.drawImage(ImageReader(img), x0, baseline_y, width=w_pt, height=h_pt, mask="auto")

    def row_bottom(top):
        return top + 12.0

    period_top = rows["period"]
    period_bottom = row_bottom(period_top)
    cover(122.7, period_top, period_bottom, x1=183)
    draw_text(122.7, period_bottom - 0.05, data["start_date"])
    cover(254.7, period_top, period_bottom, x1=315)
    draw_text(254.7, period_bottom - 0.05, data["end_date"])

    store_top = rows["store_name"]
    store_bottom = row_bottom(store_top)
    cover(122.7, store_top, store_bottom)
    draw_text(122.7, store_bottom - 0.05, data["store_name"])

    addr_top = rows["address"]
    addr_bottom = row_bottom(addr_top)
    cover(122.7, addr_top, addr_bottom)
    draw_text(122.7, addr_bottom - 0.05, data["address"])

    # 재물부문(재고/시설비품/건물) 항목별로, 정산양식에 금액이 기재된 항목만 골라서 나열함
    # (예: 건물 값이 없으면 '건물' 자체를 언급하지 않고 재고/시설비품만 표시). 셋 다 없으면
    # '보상한도액' 줄 자체를(라벨 포함) 지움.
    limit_top = rows["limit"]
    limit_bottom = row_bottom(limit_top)
    property_items = []
    if data.get("stock_amt", 0):
        property_items.append(f'재고 {data["stock_amt"]:,}원')
    if data.get("facility_amt", 0):
        property_items.append(f'시설/비품 {data["facility_amt"]:,}원')
    if data.get("building_amt", 0):
        property_items.append(f'건물 {data["building_amt"]:,}원')
    if property_items:
        cover(188.7, limit_top, limit_bottom)  # '재고'/'시설/비품' 등 기존 라벨+금액을 통째로 지우고
        draw_text(188.7, limit_bottom - 0.05, ", ".join(property_items))  # 있는 항목만 새로 나열
    else:
        cover(56.7, limit_top, limit_bottom)  # '보상한도액 :' 라벨까지 통째로 지움

    # 영업배상 부문(평수 기준 산정)이 정산양식에 기재 안 돼있으면, 서식에 고정으로 박혀 있는
    # '배상부문 ...' / '구내치료비 ...' 두 줄을 흰색으로 지움(이 두 줄은 모든 브랜드가 항상
    # 재물부문 줄에서 정확히 +18pt, +36pt 아래에 있어서 좌표를 새로 재지 않아도 계산 가능함).
    if not data.get("has_liability", True):
        liability_row1 = limit_top + 18.0
        liability_row2 = limit_top + 36.3
        cover(122.7, liability_row1, liability_row1 + 18.3)
        cover(122.7, liability_row2, liability_row2 + 18.3)

    premium_top = rows["premium"]
    premium_bottom = row_bottom(premium_top)
    cover(122.7, premium_top, premium_bottom, x1=rows.get("premium_x1", 160))
    draw_text(122.7, premium_bottom - 0.05, f'{data["premium"]:,}')

    c.save()
    buf.seek(0)

    overlay_reader = PdfReader(buf)
    template_reader = PdfReader(_cert_template_path_for_brand(brand))
    writer = PdfWriter()
    page = template_reader.pages[0]
    page.merge_page(overlay_reader.pages[0])
    writer.add_page(page)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


# 브랜드/담당자마다 같은 항목을 다른 이름으로 적는 경우가 있어서, 정규화 후 이 표로 한 번 더 통일함
# (예: 올리비아로렌 취합본은 '재고' 대신 '재고자산'이라고 씀 — 이걸 못 맞추면 재고 금액이 통째로 빠짐)
HEADER_SYNONYMS = {
    "재고자산": "재고",
    "시설": "시설/비품",
}


def _norm_header(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\n", "").strip()
    s = re.sub(r"\([^)]*\)", "", s)  # 괄호 안 요율/단가 등은 파일마다 달라서 제거
    s = s.replace(" ", "")
    return HEADER_SYNONYMS.get(s, s)


def _find_header_row(ws, search_upto: int = 6) -> int | None:
    """'매장명' 라벨이 있는 행(하위 헤더 행)을 찾음. 파일마다 헤더가 2~3행 또는 3~4행일 수 있어 위치를 직접 스캔."""
    for r in range(1, min(ws.max_row, search_upto) + 1):
        row_vals = [_norm_header(c.value) for c in ws[r]]
        if "매장명" in row_vals:
            return r
    return None


_STORE_CODE_RE = re.compile(r"^[A-Za-z]{2,4}\d{2,5}[A-Za-z]?$")
_KOREAN_RE = re.compile(r"[가-힣]")


def _looks_like_store_code(v) -> bool:
    return bool(_STORE_CODE_RE.match(str(v).strip()))


def _looks_like_korean_text(v) -> bool:
    return bool(_KOREAN_RE.search(str(v)))


def _is_placeholder_code(code) -> bool:
    """'미개설(추후전달)', '코드미개설상태'처럼 매장코드가 아직 배정되지 않아서 담당자가
    임시로 적어넣은 안내 문구인 경우를 가려냄. 이런 값을 진짜 매장코드처럼 dedup 키에
    그대로 쓰면, 나중에 실제 코드가 배정된 뒤 같은 매장이 다시 파일에 실려 오거나(담당자가
    예전 이력을 안 지우고 계속 붙여보내는 경우가 있음) 문구가 미세하게 달라지면(예: '추후
    전달' vs '추후전달') 매번 새 매장으로 잘못 인식해서 중복 등록/가입증명서 재발급으로
    이어짐. 매장코드 칸의 각 줄이 전부 실제 코드 패턴(_STORE_CODE_RE)이어야 '진짜 코드'로
    인정하고, 하나라도 아니면 placeholder로 봄."""
    if not code:
        return True
    segments = [s.strip() for s in str(code).split("\n") if s.strip()]
    if not segments:
        return True
    return not all(_looks_like_store_code(seg) for seg in segments)


def _store_identity_key(vals: dict) -> tuple:
    """매장코드가 없어도 같은 매장인지 알아볼 수 있도록, 매장명+사업자번호(없으면 주소)로
    만드는 보조 식별 키. 공백/줄바꿈 차이는 무시함."""
    name = re.sub(r"\s+", "", str(vals.get("매장명") or ""))
    biz_no = re.sub(r"\s+", "", str(vals.get("사업자번호") or ""))
    addr = re.sub(r"\s+", "", str(vals.get("매장주소") or ""))
    return (name, biz_no or addr)


def _build_header_map(ws) -> tuple[dict, int]:
    """헤더 행을 자동으로 찾아 '정규화된 헤더명 -> 0-based 컬럼 인덱스' 매핑과 데이터 시작 행을 반환.
    '매장명' 등 주요 라벨이 있는 행(main_row) 바로 아래 행에 병합된 그룹(재물부문 등)의
    하위 항목명(재고/시설/비품/건물 등)이 있으므로 두 행을 합쳐서 매핑을 만들고,
    실제 데이터는 그 두 행 다음부터 시작함."""
    main_row = _find_header_row(ws)
    if main_row is None:
        return {}, 4

    sub_row = main_row + 1
    row_main = [c.value for c in ws[main_row]]
    row_sub = [c.value for c in ws[sub_row]] if sub_row <= ws.max_row else []
    max_col = max(len(row_main), len(row_sub))

    header_map = {}
    for idx in range(max_col):
        vmain = row_main[idx] if idx < len(row_main) else None
        vsub = row_sub[idx] if idx < len(row_sub) else None
        # 병합된 그룹 칸(예: '재물부문(요율 0.0665%)')은 main_row에, 그 아래 세부 항목명
        # (재고/시설/비품/건물 등)은 sub_row에 있으므로 더 구체적인 sub_row를 우선함
        label = vsub if vsub not in (None, "") else vmain
        norm = _norm_header(label)
        if norm and norm not in header_map:
            header_map[norm] = idx

    data_start = sub_row + 1

    # 일부 브랜드의 '상설' 시트처럼, 헤더 글자는 '매장명'/'매장코드' 순서대로 적혀 있는데
    # 실제 데이터는 그 반대로 입력돼 있는 경우가 실제로 있었음(담당자 쪽 서식 실수). 이걸 그대로
    # 믿으면 가입증명서에 매장코드와 매장명이 뒤바뀌어 나가는 심각한 오류로 이어지므로, 실제
    # 데이터 몇 줄의 값 형태(코드형 vs 한글명)를 보고 뒤바뀐 게 확실하면 자동으로 바로잡음.
    name_idx = header_map.get("매장명")
    code_idx = header_map.get("매장코드")
    if name_idx is not None and code_idx is not None and name_idx != code_idx:
        checked = 0
        swapped_looking = 0
        for r in range(data_start, min(ws.max_row, data_start + 30) + 1):
            name_val = ws.cell(row=r, column=name_idx + 1).value
            code_val = ws.cell(row=r, column=code_idx + 1).value
            if not name_val or not code_val:
                continue
            checked += 1
            if _looks_like_store_code(name_val) and _looks_like_korean_text(code_val):
                swapped_looking += 1
            if checked >= 5:
                break
        if checked and swapped_looking == checked:
            header_map["매장명"], header_map["매장코드"] = code_idx, name_idx

    return header_map, data_start


_SIMPLE_REF_RE = re.compile(r"^=([^!]+)!(\$?[A-Z]+\$?\d+)$")


def _resolve_simple_formula(formula, wb_values):
    """'=최초가입전체리스트!E108'처럼 다른 시트의 셀 하나를 그대로 참조하는 단순 수식이면,
    그 시트의 실제 값을 대신 찾아서 반환함. 매장명 칸이 이런 수식으로 돼 있는 폐점매장 행이
    있는데, 파일이 한 번도 진짜 엑셀에서 재계산/저장된 적이 없으면 캐시된 값이 없어서
    data_only=True로 읽어도 매장명이 빈 값(None)으로 나옴 -> 중복 확인(dedup)이
    그 행을 못 찾아서 같은 매장이 두 번 추가되는 버그로 이어짐. 이걸 막기 위한 보강."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    m = _SIMPLE_REF_RE.match(formula.strip())
    if not m:
        return None
    sheet_name = m.group(1).strip().strip("'")
    cell_ref = m.group(2).replace("$", "")
    if sheet_name not in wb_values.sheetnames:
        return None
    try:
        return wb_values[sheet_name][cell_ref].value
    except Exception:
        return None


def _row_values(ws, header_map: dict, r: int, formula_ws=None, wb_values=None) -> dict:
    row_cells = ws[r]
    formula_cells = formula_ws[r] if formula_ws is not None else None
    vals = {}
    for key, idx in header_map.items():
        v = row_cells[idx].value if idx < len(row_cells) else None
        if v is None and formula_cells is not None and wb_values is not None and idx < len(formula_cells):
            resolved = _resolve_simple_formula(formula_cells[idx].value, wb_values)
            if resolved is not None:
                v = resolved
        vals[key] = v
    return vals


def _extract_data_rows(ws, header_map: dict, min_row: int, formula_ws=None, wb_values=None) -> list:
    out = []
    for r in range(min_row, ws.max_row + 1):
        vals = _row_values(ws, header_map, r, formula_ws, wb_values)
        if not vals.get("매장명"):
            continue
        vals["_excel_row"] = r
        out.append(vals)
    return out


def _val_key(v) -> str:
    """셀 값을 dedup 비교용 문자열로 정규화 (datetime이든 문자열이든 동일하게 비교 가능하게)"""
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v).strip()


def _dedup_key(vals: dict) -> tuple:
    """매장코드가 같아도 보험시작일이 다르면(단기 특판 연장 등) 별개 건으로 취급.
    보험시작일이 없으면 접수일자로 대체."""
    code = vals.get("매장코드")
    period_key = _val_key(vals.get("보험시작일")) or _val_key(vals.get("접수일자"))
    if code:
        return ("code", str(code).strip(), period_key)
    name = str(vals.get("매장명") or "").strip()
    return ("namedate", name, period_key)


def _get_brand_name(wb) -> str | None:
    """첫 시트에서 '*. 브랜드명' 형식의 셀을 찾아 브랜드명을 반환.
    보통 A1 셀에 있지만, 담당자마다 앞에 빈 열을 하나 더 두는 등 서식이 조금씩 달라서
    B1처럼 다른 칸에 적혀 있는 경우도 실제로 있었음. 그래서 A1만 보지 않고 첫 시트의
    맨 위 몇 행/열을 넓게 훑어서 '*.'로 시작하는 칸을 찾음."""
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=6):
        for cell in row:
            v = cell.value
            if not v:
                continue
            m = re.match(r"\*\.\s*(.+)", str(v).strip())
            if m:
                return m.group(1).strip()
    return None


def _find_type_sheets(wb, keyword: str) -> dict:
    """시트 이름에 keyword(예: '신규매장')가 포함된 시트를 '정상'/'상설' 등으로 분류.
    담당자가 개별로 보내는 신청서는 '신규매장(정상)'이 아니라 그냥 '신규매장'처럼
    괄호 구분 없이 오는 경우가 많은데, 이때는 일반(정상) 매장 신청으로 간주함."""
    result = {}
    for name in wb.sheetnames:
        if keyword not in name:
            continue
        if "상설" in name:
            result["상설"] = wb[name]
        else:
            result.setdefault("정상", wb[name])
    return result


def _strip_external_links(wb) -> None:
    """일부 담당자 취합본에는 예전에 다른 파일(예: 카카오톡으로 받은 엑셀)을 참조하던 '외부 워크북
    링크'가 남아있는 경우가 있음(수식이 [1]시트명! 형태). openpyxl로 다시 저장하면 이 외부 링크의
    내부 참조가 깨져서 엑셀에서 '내용에 문제가 있습니다' 복구 경고가 뜨는 원인이 됨. 다행히 같은
    이름의 시트가 통합문서 안에 이미 있는 경우가 대부분이라, [숫자] 표시만 지워서 로컬 시트를
    가리키도록 바꾸고, 더 이상 쓰이지 않는 외부 링크 정의 자체도 제거함."""
    if not getattr(wb, "_external_links", None):
        return
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and re.search(r"\[\d+\]", cell.value):
                    cell.value = re.sub(r"\[\d+\]", "", cell.value)
    wb._external_links = []


SKIP_WRITE_FIELDS = {"월별", "순번"}


def _find_target_row(ws, header_map: dict, min_row: int) -> int:
    """실제 데이터(매장명)가 채워진 가장 마지막 행 바로 다음 행을 찾음. 새 매장은 항상 접수
    순서대로 맨 아래에 쌓여야 하므로, 중간에 (예전에 지워졌거나 서식만 남은) 빈 행이 있어도
    거기 끼워넣지 않고 항상 전체에서 가장 마지막에 채워진 행 다음에 추가함."""
    name_idx = header_map.get("매장명")
    if name_idx is None:
        return ws.max_row + 1
    last_filled = min_row - 1
    for r in range(min_row, ws.max_row + 1):
        row_cells = ws[r]
        if name_idx < len(row_cells) and row_cells[name_idx].value:
            last_filled = r
    return last_filled + 1


def _find_template_row(ws, header_map: dict, min_row: int, target_row: int) -> int:
    """target_row 바로 위에서 실제 데이터(매장명)가 채워진 가장 가까운 행을 찾아 수식 기준 행으로 씀.
    (일부 파일은 빈 줄마다 수식이 끝까지 미리 채워져 있지 않고 중간에 끊겨 있어서,
    바로 위 빈 줄이 아니라 '가장 최근 실제로 채워진 행'을 기준으로 삼아야 수식이 안전하게 이어짐)"""
    name_idx = header_map.get("매장명")
    r = target_row - 1
    while r >= min_row:
        if name_idx is None or ws.cell(row=r, column=name_idx + 1).value:
            return r
        r -= 1
    return target_row


YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

# 통합파일에 새로 써넣는 날짜 칸들. 담당자가 접수일자를 '26.07.16'처럼 문자로 직접 타이핑해서
# 보내는 경우가 있어서, 그대로 쓰면 셀마다 표기 형식이 제각각이 됨. 항상 0000-00-00 형식으로
# 통일해서 보이도록, 실제 날짜값으로 바꿔서 쓰고 셀 서식도 맞춰줌.
DATE_FIELDS = {"접수일자", "보험시작일", "보험종기일", "폐점일"}


def _write_row(ws, header_map: dict, target_row: int, values: dict, template_row: int) -> None:
    """각 컬럼마다 '기준 행(template_row)'의 셀이 수식이면 그 수식을 target_row로 복사(번역)하고,
    수식이 아니면(원본 입력값 칸이면) values의 값을 그대로 채워 넣음. 필드명을 하드코딩해서
    구분하지 않고, 실제로 그 칸이 수식인지 아닌지를 셀 단위로 직접 확인하기 때문에 파일마다
    수식이 있는 칸이 달라도(예: 폐점매장의 매장명이 수식인 경우 등) 안전하게 동작함.
    (음영 처리는 이 함수가 아니라 _recolor_by_period가 접수일자 기준으로 별도로 담당함)"""
    for key, idx in header_map.items():
        if key in SKIP_WRITE_FIELDS:
            continue
        col = idx + 1
        target_cell = ws.cell(row=target_row, column=col)
        if isinstance(target_cell, MergedCell):
            continue

        template_cell = ws.cell(row=template_row, column=col)
        template_val = template_cell.value
        is_formula = isinstance(template_val, str) and template_val.startswith("=")

        if is_formula:
            if target_row != template_row:
                target_cell.value = Translator(
                    template_val, origin=template_cell.coordinate
                ).translate_formula(target_cell.coordinate)
            continue

        if key in values and values[key] not in (None, ""):
            v = values[key]
            if key in DATE_FIELDS:
                parsed = _parse_date_val(v)
                if parsed is not None:
                    v = parsed
                    target_cell.number_format = "yyyy-mm-dd"
            target_cell.value = v

    # 새로 만든 행(기존 서식 범위를 벗어난 경우)엔 순번을 이어서 채워줌
    if target_row != template_row:
        seq_idx = header_map.get("순번")
        if seq_idx is not None:
            prev = ws.cell(row=template_row, column=seq_idx + 1).value
            if isinstance(prev, (int, float)):
                ws.cell(row=target_row, column=seq_idx + 1).value = int(prev) + 1


def _current_period(today: dt.date | None = None) -> tuple[dt.date, dt.date]:
    """'해당월 16일 ~ 다음달 15일'을 하나의 정산 기간으로 봄(예: 7/27이면 7/16~8/15).
    오늘이 1~15일 사이면 아직 저번 기간(전달 16일~이번달 15일)이 진행 중인 걸로 봄."""
    today = today or dt.date.today()
    if today.day >= 16:
        start = dt.date(today.year, today.month, 16)
        if today.month == 12:
            end = dt.date(today.year + 1, 1, 15)
        else:
            end = dt.date(today.year, today.month + 1, 15)
    else:
        end = dt.date(today.year, today.month, 15)
        if today.month == 1:
            start = dt.date(today.year - 1, 12, 16)
        else:
            start = dt.date(today.year, today.month - 1, 16)
    return start, end


def _format_period_label(period_start: dt.date, period_end: dt.date) -> str:
    """정산기간을 통합파일에 실제로 쓰여있는 '26년\\n7월16일~\\n8월15일' 형식의 라벨 문자열로
    만듦. 연도가 바뀌는 기간(예: 12월16일~1월15일)은 두 연도를 다 표시함."""
    start_yy = str(period_start.year)[2:]
    if period_start.year == period_end.year:
        return f"{start_yy}년\n{period_start.month}월{period_start.day}일~\n{period_end.month}월{period_end.day}일"
    end_yy = str(period_end.year)[2:]
    return (
        f"{start_yy}년\n{period_start.month}월{period_start.day}일~\n"
        f"{end_yy}년\n{period_end.month}월{period_end.day}일"
    )


def _col_a_label_blocks(ws) -> list[tuple[str, int, int]]:
    """컬럼 A(월별)에 있는 정산기간 라벨 칸들을 전부 찾아 (라벨텍스트, 시작행, 끝행) 목록으로
    반환함. 여러 행이 병합되어 하나의 라벨을 이루는 경우와, 병합 없이 단독으로 있는 라벨(예:
    그 기간에 항목이 하나뿐인 경우) 둘 다 포함함."""
    merge_end = {mc.min_row: mc.max_row for mc in ws.merged_cells.ranges if mc.min_col == 1 and mc.max_col == 1}
    blocks = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        if isinstance(cell, MergedCell):
            continue
        v = cell.value
        if v in (None, ""):
            continue
        blocks.append((str(v), r, merge_end.get(r, r)))
    return blocks


ROLLOVER_BUFFER_ROWS = 25


def _rollover_period_block(ws, header_map: dict, target_row: int, period_label: str) -> int:
    """새 항목을 target_row에 넣기 전에 호출함. 이 항목이 속한 정산기간(period_label)에 맞는
    라벨 블록으로 target_row를 맞춰줌.
    - 통합파일은 보통 몇 달치 정산기간 블록(라벨 + 여분 행)을 미리 만들어두는 서식이라, 새
      기간으로 넘어갈 때 그 미리 만들어둔 블록까지 건너뛰어 들어가야 하는데, 지금까지는 항상
      '마지막으로 채워진 행 바로 다음'에만 넣어서 새 기간 블록을 건너뛰고 예전 기간 라벨
      아래에 잘못 들어가는 문제가 있었음.
    - 방금 끝난 기간 블록에 남아있던 미사용 여분 행은 삭제하고, 그만큼 아래 블록들을 앞으로
      당겨서 서식이 계속 깔끔하게 이어지도록 함.
    - 컬럼 A에 정산기간 라벨이 아예 없는(이 서식을 안 쓰는) 시트는 손대지 않고 그대로 둠."""
    name_idx = header_map.get("매장명")
    if name_idx is None:
        return target_row

    blocks = _col_a_label_blocks(ws)
    if not blocks:
        return target_row  # 이 서식(기간 블록 미리 만들어두기)을 안 쓰는 시트는 그대로 둠

    for label, start, end in blocks:
        if start <= target_row <= end:
            if label == period_label:
                return target_row  # 이미 맞는 기간 블록 안에 있음
            break  # target_row가 속한 블록이 있는데 기간이 다르면 아래에서 처리

    match = next(((start, end) for label, start, end in blocks if label == period_label), None)

    if match is None:
        # 미리 만들어둔 블록이 없음(향후 기간이 다 소진됨) -> 새로 라벨 + 여분 행을 만듦
        ws.cell(row=target_row, column=1).value = period_label
        end_row = target_row + ROLLOVER_BUFFER_ROWS - 1
        if end_row > target_row:
            ws.merge_cells(f"A{target_row}:A{end_row}")
        return target_row

    block_start, _block_end = match
    if block_start == target_row:
        return target_row
    if block_start < target_row:
        return target_row  # 이상 상황(라벨 블록이 이미 지나간 위치) - 안전하게 그대로 둠

    gap = block_start - target_row
    _delete_blank_rows_and_shift(ws, target_row, gap)
    return target_row


def _delete_blank_rows_and_shift(ws, start_row: int, count: int) -> None:
    """start_row부터 count개의 빈 여분 행을 지우고 그 아래 모든 행을 count칸 위로 당김.
    수식은 Translator로 새 위치에 맞게 다시 옮기고, 컬럼 A(월별)의 정산기간 라벨 병합도
    같이 당김. start_row~start_row+count-1 구간은 실제 데이터가 없는 빈 행이어야 안전함
    (호출하는 쪽인 _rollover_period_block에서 이미 그렇게 보장함)."""
    max_row = ws.max_row
    max_col = ws.max_column

    col_a_merges = [mc for mc in list(ws.merged_cells.ranges) if mc.min_col == 1 and mc.max_col == 1]

    to_shift = []
    for mc in col_a_merges:
        if mc.min_row >= start_row:
            to_shift.append(mc)
        elif mc.max_row >= start_row:
            # 방금 끝난 기간의 라벨 블록이 삭제 구간까지 넘어와 있던 경우 -> 실제 데이터가
            # 있는 부분(삭제 구간 시작 직전)까지로 줄임(당기지는 않음)
            ws.unmerge_cells(str(mc))
            new_end = start_row - 1
            if new_end > mc.min_row:
                ws.merge_cells(f"A{mc.min_row}:A{new_end}")

    for mc in to_shift:
        ws.unmerge_cells(str(mc))

    for r in range(start_row + count, max_row + 1):
        dst_row = r - count
        for c in range(1, max_col + 1):
            src = ws.cell(row=r, column=c)
            dst = ws.cell(row=dst_row, column=c)
            if isinstance(dst, MergedCell):
                continue
            val = None if isinstance(src, MergedCell) else src.value
            if isinstance(val, str) and val.startswith("="):
                try:
                    val = Translator(val, origin=src.coordinate).translate_formula(dst.coordinate)
                except Exception:
                    pass
            dst.value = val
            if not isinstance(src, MergedCell):
                dst.number_format = src.number_format
                dst.fill = copy.copy(src.fill)
                dst.font = copy.copy(src.font)
                dst.alignment = copy.copy(src.alignment)
                dst.border = copy.copy(src.border)

    for r in range(max(max_row - count + 1, start_row), max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    for mc in to_shift:
        new_min = mc.min_row - count
        new_max = mc.max_row - count
        if new_max >= new_min and new_min >= 1:
            ws.merge_cells(f"A{new_min}:A{new_max}")

    if max_row - count >= start_row:
        try:
            ws.delete_rows(max_row - count + 1, count)
        except Exception:
            pass


def _parse_date_val(date_val) -> dt.date | None:
    """접수일자 셀 값을 날짜로 변환. datetime/date 객체는 그대로, 문자열('26.07.16',
    '2026-07-16', '2026.07.16' 등)은 여러 형식을 시도해서 파싱. 담당자가 개별로 셀에
    직접 타이핑한 파일은 날짜가 진짜 날짜형이 아니라 문자열로 들어오는 경우가 있음."""
    if date_val is None:
        return None
    if hasattr(date_val, "date"):
        return date_val.date()
    if isinstance(date_val, dt.date):
        return date_val
    s = str(date_val).strip()
    if not s:
        return None
    for fmt in ("%y.%m.%d", "%Y.%m.%d", "%y-%m-%d", "%Y-%m-%d", "%y/%m/%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _is_stale_recv_date(vals: dict, period_start: dt.date) -> str | None:
    """담당자가 보내는 정산엑셀에는 이번에 새로 접수된 건뿐 아니라 그동안의 전체 이력이 통째로
    들어있는 경우가 많음. 그중 매장코드가 통합파일에 없는 행은 규칙상 '신규'로 처리하지만,
    접수일자가 지난 정산기간보다도 훨씬 이전(한 기간 이상 전)이면 사실은 예전에 이미 다른 방식으로
    처리된 건이 파일에만 남아있는 것일 수 있음. 이런 경우를 조용히 신규 처리하지 말고 사용자에게
    표시해서 실수로 잘못 등록되는 걸 막기 위한 확인용 함수. 오래된 날짜면 'YYYY-MM-DD' 문자열을,
    아니면 None을 반환함."""
    recv = vals.get("접수일자") or vals.get("보험시작일")
    d = _parse_date_val(recv)
    if d is None:
        return None
    threshold = period_start - dt.timedelta(days=31)
    if d < threshold:
        return d.isoformat()
    return None


# 음영(fill) 처리에서만 예외로 둘 칸. 값은 절대 건드리면 안 되는 '월별'(정산기간 라벨, 여러
# 행에 병합돼 있어 그 블록 전체를 대표하는 텍스트라 행 단위로 색칠하면 오히려 어색함)만 제외하고,
# '순번'은 _write_row에서는 값을 안 건드릴 뿐 음영까지 뺄 이유는 없음 - 예전엔 SKIP_WRITE_FIELDS를
# 그대로 재사용해서 순번 칸 음영이 기간이 지나도 노란색으로 남아있는 버그가 있었음.
SKIP_FILL_FIELDS = {"월별"}


def _recolor_by_period(ws, header_map: dict, min_row: int, period_start: dt.date, period_end: dt.date) -> None:
    """접수일자가 이번 정산 기간(16일~다음달 15일) 안이면 노란색, 기간이 지난 항목이면 다시
    흰색(음영 없음)으로 되돌림. 월별 칸(정산기간 라벨)만 그대로 두고, 순번을 포함한 나머지
    칸은 전부 같이 칠함. 신규매장/폐점매장 시트 양쪽 다 이 규칙을 적용해서, 이번 달에 새로
    들어온 신규/폐점 매장을 한눈에 볼 수 있게 함."""
    date_idx = header_map.get("접수일자")
    name_idx = header_map.get("매장명")
    if date_idx is None or name_idx is None:
        return
    for r in range(min_row, ws.max_row + 1):
        name_val = ws.cell(row=r, column=name_idx + 1).value
        if not name_val:
            continue
        date_val = ws.cell(row=r, column=date_idx + 1).value
        row_date = _parse_date_val(date_val)
        in_period = row_date is not None and period_start <= row_date <= period_end
        fill = YELLOW_FILL if in_period else NO_FILL
        for key, idx in header_map.items():
            if key in SKIP_FILL_FIELDS:
                continue
            cell = ws.cell(row=r, column=idx + 1)
            if isinstance(cell, MergedCell):
                continue
            cell.fill = fill


def _extract_rate(ws, header_map: dict, field: str, pattern: str, default: float, min_row: int) -> float:
    idx = header_map.get(field)
    if idx is None:
        return default
    for r in range(min_row, min(ws.max_row, min_row + 50) + 1):
        cell = ws.cell(row=r, column=idx + 1)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            m = re.search(pattern, cell.value)
            if m:
                return float(m.group(1))
    return default


def _compute_new_store_cert_values(vals: dict, rate1_pct: float, rate2: float) -> dict:
    stock = float(vals.get("재고") or 0)
    facility = float(vals.get("시설/비품") or 0)
    building = float(vals.get("건물") or 0)
    total = stock + facility + building
    property_premium = total * rate1_pct / 100
    pyeong = float(vals.get("평수") or 0)
    liability_premium = pyeong * rate2
    start = vals.get("보험시작일")
    end = vals.get("보험종기일")
    if hasattr(start, "date"):
        start_d = start
    else:
        start_d = None
    if hasattr(end, "date"):
        end_d = end
    else:
        end_d = None
    if start_d and end_d:
        days = (end_d - start_d).days
        premium = (property_premium + liability_premium) * days / 365
    else:
        premium = 0
    return {
        "stock_amt": int(stock),
        "facility_amt": int(facility),
        "building_amt": int(building),
        "premium": round(premium),
        "start_date": start_d.strftime("%Y.%m.%d") if start_d else str(start or ""),
        "start_date_yymmdd": start_d.strftime("%y%m%d") if start_d else "",
        "end_date": end_d.strftime("%Y.%m.%d") if end_d else str(end or ""),
        # 재고/시설비품/건물 중 하나라도 기재돼 있으면 재물부문 있음, 평수가 기재돼 있으면 영업배상 부문 있음.
        # 둘 다 없으면 가입증명서에서 해당 줄 자체를 지워야 함(사용자 요청).
        "has_property": bool(stock or facility or building),
        "has_liability": bool(pyeong),
        # 보험종기일이 이미 지난(오늘 이전) 건은 통합파일엔 기록해두되, 가입증명서는 만들 필요
        # 없음(사용자 요청) - 예: 뒤늦게 딸려온, 이미 끝난 단기 행사장 계약 등
        "period_ended": bool(end_d and end_d.date() < dt.date.today()),
    }


def _load_policy_numbers() -> dict:
    try:
        with open(POLICY_NUMBERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_policy_numbers(policy_numbers: dict) -> None:
    os.makedirs(os.path.dirname(POLICY_NUMBERS_FILE), exist_ok=True)
    with open(POLICY_NUMBERS_FILE, "w", encoding="utf-8") as f:
        json.dump(policy_numbers, f, ensure_ascii=False, indent=2)


async def set_policy_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "사용법: /setpolicy <브랜드명> <증권번호>\n"
            "예: /setpolicy 트레몰로 82509565736000\n\n"
            "브랜드명은 엑셀 시트 A1 셀에 적힌 이름(예: '트레몰로', '월메이드')과 정확히 같아야 해요."
        )
        return
    brand = context.args[0]
    policy_no = context.args[1]
    policy_numbers = _load_policy_numbers()
    policy_numbers[brand] = policy_no
    _save_policy_numbers(policy_numbers)
    await update.message.reply_text(f"✅ '{brand}' 브랜드의 증권번호를 등록했어요: {policy_no}")


def _load_contacts() -> dict:
    try:
        with open(CONTACTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_contacts(contacts: dict) -> None:
    os.makedirs(os.path.dirname(CONTACTS_FILE), exist_ok=True)
    with open(CONTACTS_FILE, "w", encoding="utf-8") as f:
        json.dump(contacts, f, ensure_ascii=False, indent=2)


def _norm_contact_name(name: str) -> str:
    return re.sub(r"\s+", "", str(name or "")).strip()


async def set_contact_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "사용법: /setcontact <담당자 이름> <이메일>\n"
            "예: /setcontact 안해인 hian01@sejung.co.kr\n\n"
            "이름은 정산양식 엑셀의 '접수자' 칸에 적힌 이름과 정확히 같아야 해요. "
            "등록해두면, 그 담당자가 텔레그램으로 직접 올린 정산양식에서도 신규매장 "
            "가입증명서를 자동으로 그 담당자 이메일로 보내드려요."
        )
        return
    name = context.args[0]
    email_addr = context.args[1]
    contacts = _load_contacts()
    contacts[_norm_contact_name(name)] = email_addr
    _save_contacts(contacts)
    await update.message.reply_text(f"✅ '{name}' 담당자의 이메일을 등록했어요: {email_addr}")


async def list_contacts_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    contacts = _load_contacts()
    if not contacts:
        await update.message.reply_text("등록된 담당자 연락처가 없어요. /setcontact로 등록해주세요.")
        return
    lines = "\n".join(f"- {name}: {addr}" for name, addr in contacts.items())
    await update.message.reply_text(f"등록된 담당자 연락처:\n{lines}")


def _load_pending_certs() -> dict:
    try:
        with open(PENDING_CERTS_INDEX_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_pending_certs(pending: dict) -> None:
    os.makedirs(os.path.dirname(PENDING_CERTS_INDEX_FILE), exist_ok=True)
    with open(PENDING_CERTS_INDEX_FILE, "w", encoding="utf-8") as f:
        json.dump(pending, f, ensure_ascii=False, indent=2)


def _queue_pending_cert(store_name: str, target_email: str, subject: str, body: str, out_name: str, pdf_bytes: bytes) -> str:
    """가입증명서를 바로 보내지 않고 대기시켜둠. 유진님이 '보내기'를 누르면 그때 실제 발송함."""
    os.makedirs(PENDING_CERTS_DIR, exist_ok=True)
    cert_id = uuid.uuid4().hex[:10]
    pdf_path = os.path.join(PENDING_CERTS_DIR, f"{cert_id}.pdf")
    with open(pdf_path, "wb") as f:
        f.write(pdf_bytes)
    pending = _load_pending_certs()
    pending[cert_id] = {
        "store_name": store_name,
        "target_email": target_email,
        "subject": subject,
        "body": body,
        "out_name": out_name,
        "pdf_path": pdf_path,
    }
    _save_pending_certs(pending)
    return cert_id


def _remove_pending_cert(cert_id: str) -> dict | None:
    pending = _load_pending_certs()
    entry = pending.pop(cert_id, None)
    if entry is None:
        return None
    _save_pending_certs(pending)
    try:
        os.remove(entry["pdf_path"])
    except OSError:
        pass
    return entry


async def handle_cert_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if query is None:
        return
    if query.from_user is None or query.from_user.id != ALLOWED_USER_ID:
        await query.answer()
        return

    data = query.data or ""
    if ":" not in data:
        await query.answer()
        return
    action, cert_id = data.split(":", 1)

    pending = _load_pending_certs()
    entry = pending.get(cert_id)
    if entry is None:
        await query.answer()
        await query.edit_message_text("이미 처리됐거나 없는 요청이에요.")
        return

    if action == "sendcert":
        await query.answer()
        try:
            with open(entry["pdf_path"], "rb") as f:
                pdf_bytes = f.read()
            _send_email(
                entry["target_email"],
                subject=entry["subject"],
                body=entry["body"],
                attachments=[(entry["out_name"], pdf_bytes)],
            )
            _remove_pending_cert(cert_id)
            await query.edit_message_text(f"✅ '{entry['store_name']}' 가입증명서를 {entry['target_email']}로 보냈어요.")
        except Exception:
            logger.exception("대기 중이던 가입증명서 발송 중 오류")
            await query.edit_message_text(f"⚠️ '{entry['store_name']}' 가입증명서 발송에 실패했어요. 이 메시지에서 다시 눌러 시도해주세요.")
    elif action == "holdcert":
        # 메시지와 버튼을 그대로 남겨둠 -> 명령어 없이도, 나중에 이 메시지를 찾아 스크롤해서
        # 그냥 다시 '보내기'를 누르기만 하면 됨. 살짝 알려주는 팝업만 띄우고 메시지는 안 건드림.
        await query.answer("⏸ 대기 상태예요. 나중에 이 메시지에서 다시 눌러 보내시면 돼요.", show_alert=True)
    elif action == "cancelcert":
        await query.answer()
        _remove_pending_cert(cert_id)
        await query.edit_message_text(f"🗑 '{entry['store_name']}' 가입증명서 발송을 취소했어요.")


async def list_pending_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    pending = _load_pending_certs()
    if not pending:
        await update.message.reply_text("대기 중인 가입증명서가 없어요.")
        return
    for cert_id, entry in pending.items():
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ 보내기", callback_data=f"sendcert:{cert_id}"),
            InlineKeyboardButton("🗑 취소", callback_data=f"cancelcert:{cert_id}"),
        ]])
        await update.message.reply_text(
            f"'{entry['store_name']}' → {entry['target_email']}",
            reply_markup=keyboard,
        )


def _load_brand_names() -> dict:
    try:
        with open(BRAND_NAMES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_brand_names(brand_names: dict) -> None:
    os.makedirs(os.path.dirname(BRAND_NAMES_FILE), exist_ok=True)
    with open(BRAND_NAMES_FILE, "w", encoding="utf-8") as f:
        json.dump(brand_names, f, ensure_ascii=False, indent=2)


def _display_name(brand: str) -> str:
    """봇이 파일을 보내드릴 때 쓸 이름. 따로 지정해둔 게 있으면 그 이름, 없으면 브랜드명(A1 셀 이름) 그대로."""
    return _load_brand_names().get(brand, brand)


async def set_brand_name_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "사용법: /setbrandname <브랜드명> <파일에 표시할 이름>\n"
            "예: /setbrandname 올리비아로렌&주얼리 올리비아로렌&오뷔엘알&주얼리\n\n"
            "브랜드명(첫 번째)은 /brands 로 확인되는 이름과 정확히 같아야 해요. "
            "엑셀 A1 셀이나 신규/폐점 매장 비교 기준은 그대로 두고, 봇이 보내드리는 통합파일의 이름만 바뀌어요."
        )
        return
    brand = context.args[0]
    new_name = " ".join(context.args[1:])
    brand_names = _load_brand_names()
    brand_names[brand] = new_name
    _save_brand_names(brand_names)
    await update.message.reply_text(f"✅ 이제 '{brand}' 통합파일을 보내드릴 때 '{new_name}.xlsx'로 보내드릴게요.")


def _load_brand_aliases() -> dict:
    try:
        with open(BRAND_ALIASES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_brand_aliases(aliases: dict) -> None:
    os.makedirs(os.path.dirname(BRAND_ALIASES_FILE), exist_ok=True)
    with open(BRAND_ALIASES_FILE, "w", encoding="utf-8") as f:
        json.dump(aliases, f, ensure_ascii=False, indent=2)


def _resolve_brand_alias(brand: str) -> str:
    """담당자마다 A1 셀에 브랜드명을 조금씩 다르게 적어 보내는 경우가 있어서(예: '올리비아로렌&주얼리'
    vs '올리비아로렌&오뷔엘알&주얼리'), 등록해둔 별칭이 있으면 진짜(정식) 브랜드명으로 바꿔줌.
    이걸 안 하면 같은 브랜드인데 이름이 살짝 달라서 별도의 새 통합파일로 쪼개져 버림."""
    return _load_brand_aliases().get(brand, brand)


async def set_brand_alias_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "사용법: /setbrandalias <다르게 인식된 이름> <진짜 브랜드명>\n"
            "예: /setbrandalias 올리비아로렌&주얼리 올리비아로렌&오뷔엘알&주얼리\n\n"
            "어떤 담당자가 보낸 엑셀의 브랜드명이 <다르게 인식된 이름>으로 나오면, 이제부터 <진짜 브랜드명>의 "
            "통합파일로 합쳐서 처리해요. 진짜 브랜드명은 /brands 로 확인할 수 있어요."
        )
        return
    alias = context.args[0]
    canonical = " ".join(context.args[1:])
    aliases = _load_brand_aliases()
    aliases[alias] = canonical
    _save_brand_aliases(aliases)
    await update.message.reply_text(f"✅ 앞으로 '{alias}'로 인식되는 엑셀은 '{canonical}' 통합파일로 합쳐서 처리할게요.")


async def set_cert_template_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """새 브랜드의 실제 발급된 예시 가입증명서 PDF를 그 브랜드의 서식으로 등록.
    사용법: 예시 PDF를 먼저 보낸 뒤, 그 PDF 메시지에 '답장(reply)'하면서 '/setcerttemplate 브랜드명'을 입력."""
    if not is_allowed(update):
        return
    if not context.args:
        await update.message.reply_text(
            "사용법: 먼저 그 브랜드로 실제 발급된 예시 가입증명서 PDF를 보내주세요. 그 다음 그 PDF 메시지에 "
            "답장(reply)하면서 아래처럼 입력해주세요.\n\n/setcerttemplate <브랜드명>\n예: /setcerttemplate 트레몰로\n\n"
            "브랜드명은 /brands 로 확인되는 이름과 정확히 같아야 해요."
        )
        return
    reply = update.message.reply_to_message
    if not reply or not reply.document:
        await update.message.reply_text(
            "가입증명서 예시 PDF 파일에 답장(reply)하는 방식으로 이 명령어를 보내주세요.\n"
            "(PDF를 먼저 보내고, 그 메시지를 길게 눌러 '답장'을 선택한 뒤 명령어를 입력)"
        )
        return
    doc = reply.document
    if not (doc.file_name or "").lower().endswith(".pdf"):
        await update.message.reply_text("PDF 파일에 답장해주세요.")
        return
    brand = " ".join(context.args)
    try:
        tg_file = await doc.get_file()
        file_bytes = bytes(await tg_file.download_as_bytearray())
    except Exception:
        logger.exception("가입증명서 서식 다운로드 중 오류")
        await update.message.reply_text("파일을 받아오는 중 오류가 발생했어요. 다시 시도해주세요.")
        return
    os.makedirs(CERT_TEMPLATES_DIR, exist_ok=True)
    with open(os.path.join(CERT_TEMPLATES_DIR, f"{brand}.pdf"), "wb") as f:
        f.write(file_bytes)
    await update.message.reply_text(
        f"✅ '{brand}' 가입증명서 서식으로 등록했어요. 이제부터 이 브랜드는 이 서식을 기준으로, "
        "보험기간/점포명/주소/보상한도액/보험료만 바뀌어서 인증서가 나가요.\n\n"
        "⚠️ 다만 이 서식은 트레몰로 서식과 항목 위치가 조금 다를 수 있어서 값이 살짝 어긋나 보일 수 있어요. "
        "그런 경우 알려주시면 위치를 정확히 맞춰서 코드에 반영해드릴게요.\n"
        "또한 이 서식은 재배포하면 사라지니, 재배포 후에는 다시 등록해주셔야 해요."
    )


async def list_brands_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if not os.path.isdir(MASTERS_DIR):
        await update.message.reply_text("등록된 브랜드가 아직 없어요.")
        return
    brands = sorted(f[:-5] for f in os.listdir(MASTERS_DIR) if f.endswith(".xlsx"))
    if not brands:
        await update.message.reply_text("등록된 브랜드가 아직 없어요.")
        return
    policy_numbers = _load_policy_numbers()
    brand_names = _load_brand_names()
    aliases = _load_brand_aliases()
    lines = []
    for b in brands:
        status = "증권번호 등록됨" if b in policy_numbers else "⚠️ 증권번호 미등록"
        if b in brand_names:
            status += f", 표시명: {brand_names[b]}"
        alias_list = [a for a, c in aliases.items() if c == b]
        if alias_list:
            status += f", 별칭: {', '.join(alias_list)}"
        lines.append(f"- {b} ({status})")
    await update.message.reply_text("📋 등록된 브랜드 목록:\n" + "\n".join(lines))


async def reset_brand_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return
    if not context.args:
        await update.message.reply_text(
            "사용법: /resetbrand <브랜드명>\n예: /resetbrand 월메이드\n\n"
            "해당 브랜드의 통합파일을 삭제해요. 다음에 그 브랜드 엑셀을 올리면 그 파일을 새 기준으로 다시 등록해요.\n"
            "정확한 브랜드명은 /brands 로 확인할 수 있어요."
        )
        return
    brand = " ".join(context.args)
    master_path = os.path.join(MASTERS_DIR, f"{brand}.xlsx")
    if os.path.exists(master_path):
        os.remove(master_path)
        await update.message.reply_text(f"🗑️ '{brand}' 통합파일을 삭제했어요. 다음에 이 브랜드 엑셀을 올리면 그 파일을 새 기준으로 등록할게요.")
    else:
        await update.message.reply_text(f"'{brand}' 통합파일을 찾지 못했어요. /brands 로 정확한 브랜드명을 확인해주세요.")


async def send_master_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """지금 저장되어 있는 통합파일을 새로 올릴 필요 없이 그대로(최신 파일명 등 반영해서) 다시 받아봄"""
    if not is_allowed(update):
        return
    if not context.args:
        await update.message.reply_text(
            "사용법: /sendmaster <브랜드명>\n예: /sendmaster 트레몰로\n\n"
            "지금 저장되어 있는 통합파일을 다시 보내드려요. 정확한 브랜드명은 /brands 로 확인할 수 있어요."
        )
        return
    brand = " ".join(context.args)
    master_path = os.path.join(MASTERS_DIR, f"{brand}.xlsx")
    if not os.path.exists(master_path):
        await update.message.reply_text(f"'{brand}' 통합파일을 찾지 못했어요. /brands 로 정확한 브랜드명을 확인해주세요.")
        return
    with open(master_path, "rb") as f:
        master_bytes = f.read()
    display_name = _display_name(brand)
    await update.message.reply_document(
        document=io.BytesIO(master_bytes),
        filename=f"{display_name}.xlsx",
        caption=f"📎 '{display_name}' 통합파일이에요.",
    )


def _sync_brand_excel(file_bytes: bytes) -> dict | None:
    """엑셀을 읽어 브랜드를 판별하고, 저장된 통합파일과 비교해 신규/폐점 매장을 반영.
    반환: {"brand", "cold_start", "new_stores": [...], "closed_count": int, "master_bytes": bytes}
    브랜드를 못 찾으면 None."""
    try:
        input_wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        logger.exception("엑셀 파일을 여는 중 오류")
        return None

    brand = _get_brand_name(input_wb)
    if not brand:
        return None
    brand = _resolve_brand_alias(brand)

    os.makedirs(MASTERS_DIR, exist_ok=True)
    master_path = os.path.join(MASTERS_DIR, f"{brand}.xlsx")

    if not os.path.exists(master_path):
        with open(master_path, "wb") as f:
            f.write(file_bytes)
        return {"brand": brand, "cold_start": True, "new_stores": [], "closed_count": 0, "master_bytes": file_bytes}

    # 쓰기용(수식 보존)과 읽기용(중복 확인은 계산된 값 기준) 두 벌로 마스터 파일을 엶.
    # 폐점매장 시트처럼 '매장명' 칸 자체가 VLOOKUP 수식인 행이 실제로 존재하기 때문에,
    # 중복 판별은 반드시 계산된 값(data_only=True) 기준으로 해야 정확함.
    master_wb = openpyxl.load_workbook(master_path, data_only=False)
    master_wb_values = openpyxl.load_workbook(master_path, data_only=True)

    policy_numbers = _load_policy_numbers()
    policy_no = policy_numbers.get(brand, "")

    # 오래된 접수일자인 '신규' 건을 나중에 알림으로 표시하기 위해, 그리고 정산기간 라벨
    # 블록을 맞춰넣기 위해 미리 계산해둠
    period_start, period_end = _current_period()
    current_period_label = _format_period_label(period_start, period_end)

    new_stores = []
    closed_count = 0

    input_new_sheets = _find_type_sheets(input_wb, "신규매장")
    master_new_sheets = _find_type_sheets(master_wb, "신규매장")
    master_new_sheets_values = _find_type_sheets(master_wb_values, "신규매장")

    # 담당자가 보내는 개별 신청서는 '신규매장(정상)/신규매장(상설)' 구분 없이 그냥 '신규매장'
    # 하나로만 오는 경우가 많아서(_find_type_sheets 설명 참고), 이때는 일단 '정상'으로 간주해서
    # 비교함. 그런데 그 안에 사실은 이미 '상설' 쪽에 등록된 매장이 섞여 들어오면, '정상' 쪽
    # 기존 목록에는 없으니 신규로 잘못 잡혀서 똑같은 매장이 중복 등록되는 사고가 남. 이를 막기
    # 위해 어떤 서브타입을 처리하든 정상/상설 등 모든 서브타입에 이미 등록된 매장 전체를 합쳐서
    # 중복 여부를 확인함(실제로 새 행을 쓰는 시트는 여전히 해당 서브타입 시트 하나뿐).
    all_new_existing_keys = set()
    all_new_existing_identities = set()
    for _st, _mws in master_new_sheets.items():
        _mws_values = master_new_sheets_values.get(_st)
        if _mws_values is None:
            continue
        _h, _mr = _build_header_map(_mws)
        if not _h:
            continue
        _existing_rows = _extract_data_rows(_mws_values, _h, _mr, formula_ws=_mws, wb_values=master_wb_values)
        all_new_existing_keys |= {_dedup_key(v) for v in _existing_rows}
        all_new_existing_identities |= {_store_identity_key(v) for v in _existing_rows}

    skipped_placeholder_stores = []

    for sub_type, input_ws in input_new_sheets.items():
        master_ws = master_new_sheets.get(sub_type)
        master_ws_values = master_new_sheets_values.get(sub_type)
        if master_ws is None or master_ws_values is None:
            continue
        input_header, input_min_row = _build_header_map(input_ws)
        master_header, master_min_row = _build_header_map(master_ws)
        if not input_header or not master_header:
            continue
        existing_keys = set(all_new_existing_keys)
        existing_identities = set(all_new_existing_identities)
        rate1 = _extract_rate(master_ws, master_header, "연간재물보험료", r"\*([\d.]+)%", 0.0665, master_min_row)
        rate2 = _extract_rate(master_ws, master_header, "연간영업배상보험료", r"\*([\d.]+)", 1793, master_min_row)

        for vals in _extract_data_rows(input_ws, input_header, input_min_row):
            # 매장코드가 아직 '미개설(추후전달)' 같은 임시 문구인 채로 왔는데, 매장명+사업자번호
            # (또는 주소)가 이미 등록된 매장과 같으면 -- 실제 코드가 나중에 배정되어 등록된 뒤
            # 담당자가 예전 임시 이력을 다시 보낸 경우일 수 있음. 이 경우 매장코드 문구 자체가
            # 매번 미세하게 달라질 수 있어(dedup 키로는 안 걸러짐) 그대로 두면 같은 매장이 계속
            # 중복 등록되므로, 여기서 별도로 걸러서 건너뜀(통합파일에 추가/증명서 발급 안 함)
            if _is_placeholder_code(vals.get("매장코드")) and _store_identity_key(vals) in existing_identities:
                skipped_placeholder_stores.append({
                    "store_name": str(vals.get("매장명") or "").strip(),
                })
                continue

            key = _dedup_key(vals)
            if key in existing_keys:
                continue
            existing_keys.add(key)
            existing_identities.add(_store_identity_key(vals))

            # 오래된 접수일자 경고는 정산파일에 실제로 적혀 있던 원래 접수일자 기준으로
            # 판단해야 하므로, 접수일자를 오늘 날짜로 덮어쓰기 전에 미리 계산해둠
            stale_recv_date = _is_stale_recv_date(vals, period_start)

            # 접수일자는 정산파일/메일에 뭐라고 적혀 있든, 실제로 봇이 이 건을 처리한 날짜
            # (메일이 온 날 또는 정산파일을 텔레그램에 업로드한 날)로 통일해서 통합파일에 기록함
            vals = dict(vals)
            vals["접수일자"] = dt.date.today()

            target_row = _find_target_row(master_ws, master_header, master_min_row)
            target_row = _rollover_period_block(master_ws, master_header, target_row, current_period_label)
            template_row = _find_template_row(master_ws, master_header, master_min_row, target_row)
            _write_row(master_ws, master_header, target_row, vals, template_row)

            cert_vals = _compute_new_store_cert_values(vals, rate1, rate2)
            address = str(vals.get("매장주소") or "").strip()
            # 엑셀 매장주소가 "...52\n(백화점내)"처럼 줄바꿈으로 나뉘어 있는 경우가 있는데,
            # 그대로 두면 가입증명서에 한 줄만 그려서 괄호 부분이 아예 안 보이게 됨(PDF 렌더링이
            # 한 줄짜리로만 처리됨). 줄바꿈을 공백으로 바꿔서 괄호 내용까지 한 줄로 전부 표시함.
            address = re.sub(r"\s*\n\s*", " ", address)
            address = re.sub(r"(?<=\S)\(", " (", address)
            new_stores.append({
                "policy_no": policy_no or DEFAULT_POLICY_NO,
                "store_code": str(vals.get("매장코드") or "").strip(),
                "store_name": str(vals.get("매장명") or "").strip(),
                "address": address,
                "received_by": str(vals.get("접수자") or "").strip(),
                "stale_recv_date": stale_recv_date,
                **cert_vals,
            })

    input_closed_sheets = _find_type_sheets(input_wb, "폐점매장")
    master_closed_sheets = _find_type_sheets(master_wb, "폐점매장")
    master_closed_sheets_values = _find_type_sheets(master_wb_values, "폐점매장")

    # 신규매장과 마찬가지로, 정상/상설 등 모든 서브타입에 이미 등록된 폐점매장을 합쳐서 중복 검사함.
    # (폐점매장 시트는 애초에 '매장코드' 칸이 없는 서식이라 미배정 코드 문제가 생기지 않으므로,
    # 신규매장과 달리 여기는 원래 이름+날짜 기준 dedup만으로 충분함)
    all_closed_existing_keys = set()
    for _st, _mws in master_closed_sheets.items():
        _mws_values = master_closed_sheets_values.get(_st)
        if _mws_values is None:
            continue
        _h, _mr = _build_header_map(_mws)
        if not _h:
            continue
        all_closed_existing_keys |= {
            _dedup_key(v)
            for v in _extract_data_rows(_mws_values, _h, _mr, formula_ws=_mws, wb_values=master_wb_values)
        }

    for sub_type, input_ws in input_closed_sheets.items():
        master_ws = master_closed_sheets.get(sub_type)
        master_ws_values = master_closed_sheets_values.get(sub_type)
        if master_ws is None or master_ws_values is None:
            continue
        input_header, input_min_row = _build_header_map(input_ws)
        master_header, master_min_row = _build_header_map(master_ws)
        if not input_header or not master_header:
            continue
        existing_keys = set(all_closed_existing_keys)

        for vals in _extract_data_rows(input_ws, input_header, input_min_row):
            key = _dedup_key(vals)
            if key in existing_keys:
                continue
            existing_keys.add(key)

            # 폐점매장도 마찬가지로 접수일자를 실제 처리일(오늘)로 통일해서 기록함
            vals = dict(vals)
            vals["접수일자"] = dt.date.today()

            target_row = _find_target_row(master_ws, master_header, master_min_row)
            target_row = _rollover_period_block(master_ws, master_header, target_row, current_period_label)
            template_row = _find_template_row(master_ws, master_header, master_min_row, target_row)
            _write_row(master_ws, master_header, target_row, vals, template_row)
            closed_count += 1

    if not new_stores and not closed_count:
        return {
            "brand": brand,
            "cold_start": False,
            "new_stores": [],
            "closed_count": 0,
            "master_bytes": None,
            "skipped_placeholder_stores": skipped_placeholder_stores,
        }

    # 이번 정산 기간(16일~다음달 15일) 기준으로 신규/폐점매장 시트 전체를 다시 칠함.
    # 새로 추가된 항목뿐 아니라, 기간이 지나 더 이상 '이번 달' 항목이 아닌 예전 노란색도
    # 흰색으로 되돌려서 항상 현재 기간만 노랗게 보이도록 함.
    period_start, period_end = _current_period()
    for sheets_dict in (master_new_sheets, master_closed_sheets):
        for master_ws in sheets_dict.values():
            header, min_row = _build_header_map(master_ws)
            if header:
                _recolor_by_period(master_ws, header, min_row, period_start, period_end)

    _strip_external_links(master_wb)

    out = io.BytesIO()
    master_wb.save(out)
    master_bytes = out.getvalue()
    with open(master_path, "wb") as f:
        f.write(master_bytes)

    return {
        "brand": brand,
        "cold_start": False,
        "new_stores": new_stores,
        "closed_count": closed_count,
        "master_bytes": master_bytes,
        "has_policy_no": bool(policy_no),
        "has_cert_template": _has_cert_template(brand),
        "skipped_placeholder_stores": skipped_placeholder_stores,
    }


async def _sync_and_notify(
    bot, chat_id: int, file_bytes: bytes, requester_email: str | None = None
) -> dict | None:
    """정산양식 엑셀을 브랜드 통합파일과 동기화하고, 결과(가입증명서/갱신된 통합파일/요약)를
    텔레그램으로 보냄. 텔레그램에 직접 파일을 올린 경우와, 메일에 첨부된 파일을 자동으로
    감지한 경우 양쪽에서 공용으로 사용함.
    브랜드를 인식하지 못하면(정산양식 형식이 아니면) 아무것도 보내지 않고 None을 반환하며,
    이 경우 어떻게 안내할지는 호출한 쪽에서 정함(메일 첨부는 정산양식이 아닐 수도 있으니
    조용히 넘어가고, 텔레그램 직접 업로드는 사용자에게 안내함).
    requester_email: 메일에 첨부되어 들어온 경우, 그 메일을 보낸 사람 주소. 신규매장
    가입증명서를 만들면 이 주소로도 자동 발송함(텔레그램으로 직접 올린 경우엔 None이라
    대신 담당자 연락처 등록부(contacts.json)에서 '접수자' 이름으로 찾음)."""
    try:
        result = _sync_brand_excel(file_bytes)
    except Exception:
        logger.exception("엑셀 동기화 중 오류")
        await bot.send_message(chat_id=chat_id, text="⚠️ 엑셀을 처리하는 중 오류가 발생했어요.")
        return None

    if result is None:
        return None

    brand = result["brand"]

    if result["cold_start"]:
        await bot.send_message(
            chat_id=chat_id,
            text=f"📁 '{brand}' 통합파일을 처음 등록했어요. 앞으로 이 파일을 기준으로 신규/폐점 매장을 비교할게요.",
        )
        return result

    skipped_placeholder_stores = result.get("skipped_placeholder_stores") or []
    if skipped_placeholder_stores:
        names = ", ".join(s["store_name"] for s in skipped_placeholder_stores)
        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"ℹ️ {names}: 매장코드가 아직 '미개설(추후전달)' 같은 임시 문구인 채로 다시 들어왔는데, "
                "매장명/사업자번호가 이미 등록된 매장과 같아서 중복 등록하지 않고 건너뛰었어요."
            ),
        )

    if not result["new_stores"] and not result["closed_count"]:
        await bot.send_message(chat_id=chat_id, text=f"'{brand}' 기준으로 새로운 신규/폐점 매장이 없어요.")
        return result

    if not result.get("has_cert_template"):
        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"⚠️ '{brand}'의 가입증명서 서식이 아직 등록되어 있지 않아, 다른 브랜드 서식으로 임시로 만들어요. "
                "증권번호/계약자 등이 실제와 다를 수 있어요. 이 브랜드로 실제 발급된 예시 인증서를 보내주시면 "
                "전용 서식으로 등록해드릴게요."
            ),
        )

    ended_stores = []
    no_contact_stores = []
    queued_stores = []
    drive_saved = 0
    contacts = _load_contacts()
    for store in result["new_stores"]:
        # 보험종기일이 이미 지난 건은 통합파일엔 기록해두되, 가입증명서는 만들지 않음(사용자 요청)
        if store.get("period_ended"):
            ended_stores.append(store)
            continue
        try:
            pdf_bytes = _build_certificate_pdf(store, brand)
        except Exception:
            logger.exception("가입증명서 생성 중 오류")
            await bot.send_message(chat_id=chat_id, text=f"⚠️ '{store['store_name']}' 가입증명서 생성에 실패했어요.")
            continue

        out_name = f"{store['store_code']}_{store['store_name']}_{store['start_date_yymmdd']}.pdf"
        await bot.send_document(
            chat_id=chat_id,
            document=io.BytesIO(pdf_bytes),
            filename=out_name,
            caption=f"📄 {store['store_name']} 가입증명서",
        )

        # 담당자에게 이메일로 보낼지: 메일로 들어온 건이면 그 발신 주소로, 텔레그램으로
        # 직접 올린 건이면 엑셀 '접수자' 이름으로 등록된 연락처를 찾음. 바로 보내지 않고
        # 유진님이 버튼으로 확인한 뒤에 실제 발송함(원치 않으면 '대기' 선택 가능)
        target_email = requester_email
        received_by = store.get("received_by") or ""
        if not target_email and received_by:
            target_email = contacts.get(_norm_contact_name(received_by))
        if target_email and GMAIL_ADDRESS and GMAIL_APP_PASSWORD:
            # 메일로 들어와서 처음 알게 된 담당자면, 다음부터 텔레그램으로 직접 올려도
            # 이메일을 찾을 수 있도록 연락처에 미리 저장해둠(발송 여부와 무관하게 저장)
            if requester_email and received_by:
                key = _norm_contact_name(received_by)
                if contacts.get(key) != requester_email:
                    contacts[key] = requester_email
                    _save_contacts(contacts)
            cert_id = _queue_pending_cert(
                store_name=store["store_name"],
                target_email=target_email,
                subject=f"[{brand}] {store['store_name']} 가입증명서",
                body=_CERT_EMAIL_BODY_TEMPLATE.format(store_name=store["store_name"]),
                out_name=out_name,
                pdf_bytes=pdf_bytes,
            )
            queued_stores.append((store["store_name"], target_email))
            keyboard = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ 보내기", callback_data=f"sendcert:{cert_id}"),
                InlineKeyboardButton("⏸ 대기", callback_data=f"holdcert:{cert_id}"),
                InlineKeyboardButton("🗑 취소", callback_data=f"cancelcert:{cert_id}"),
            ]])
            await bot.send_message(
                chat_id=chat_id,
                text=(
                    f"담당자 {received_by or target_email}님에게 '{store['store_name']}' 가입증명서를 보낼까요?\n"
                    "(대기를 누르면 이 메시지는 그대로 남아있으니, 나중에 다시 여기서 보내기를 누르시면 돼요)"
                ),
                reply_markup=keyboard,
            )
        elif received_by:
            no_contact_stores.append(received_by)

        # 구글 드라이브(설정해뒀으면 유진님 PC에 자동 동기화)에도 저장
        if _upload_to_drive(out_name, pdf_bytes):
            drive_saved += 1

    if ended_stores:
        lines = "\n".join(f"- {s['store_name']}({s['store_code']}) 종기일 {s['end_date']}" for s in ended_stores)
        await bot.send_message(
            chat_id=chat_id,
            text=(
                "ℹ️ 아래 매장은 보험종기일이 이미 지나서 통합파일에는 등록했지만 가입증명서는 만들지 않았어요.\n"
                + lines
            ),
        )

    summary = f"✅ '{brand}' 통합파일을 갱신했어요.\n신규 매장 {len(result['new_stores'])}곳"
    if result["closed_count"]:
        summary += f", 폐점 매장 {result['closed_count']}곳"
    if queued_stores:
        queued_lines = "\n".join(f"  - {name} → {addr}" for name, addr in queued_stores)
        summary += f"\n📧 담당자 발송 확인 대기 중(위 버튼 눌러주세요):\n{queued_lines}"
    if drive_saved:
        summary += f"\n💾 구글 드라이브에도 {drive_saved}건 저장했어요."
    await bot.send_message(chat_id=chat_id, text=summary)

    if no_contact_stores:
        names = ", ".join(sorted(set(no_contact_stores)))
        await bot.send_message(
            chat_id=chat_id,
            text=(
                f"ℹ️ {names} 담당자 이메일이 등록되어 있지 않아 자동 발송을 못 했어요.\n"
                "/setcontact <담당자 이름> <이메일> 로 등록해주시면 다음부터 자동으로 보내드려요."
            ),
        )

    # 담당자가 보낸 파일에 전체 이력이 섞여 있어서, 접수일자가 아주 오래된(지난 정산기간보다도
    # 이전) 매장이 '신규'로 잡힌 경우 - 이미 예전에 다른 방식으로 처리된 건일 수 있으니 등록/증명서
    # 발급은 그대로 진행하되, 놓치지 않도록 별도로 알려드림 (조용히 신규 처리만 하면 나중에 실수로
    # 발견하게 됨)
    stale_stores = [s for s in result["new_stores"] if s.get("stale_recv_date")]
    if stale_stores:
        lines = "\n".join(
            f"- {s['store_name']}({s['store_code']}) 접수일자 {s['stale_recv_date']}"
            for s in stale_stores
        )
        await bot.send_message(
            chat_id=chat_id,
            text=(
                "⚠️ 아래 매장은 접수일자가 지난 정산기간보다도 오래돼서, 이미 예전에 처리됐는데 "
                "파일에만 이력으로 남아있던 건일 수 있어요. 확인해주세요(맞으면 그냥 두시면 되고, "
                "이미 처리된 거면 알려주시면 통합파일에서 빼드릴게요).\n" + lines
            ),
        )

    if result.get("master_bytes"):
        display_name = _display_name(brand)
        await bot.send_document(
            chat_id=chat_id,
            document=io.BytesIO(result["master_bytes"]),
            filename=f"{display_name}.xlsx",
            caption=f"📎 갱신된 '{display_name}' 통합파일이에요.",
        )

    return result


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return

    draft = context.user_data.get("mail_draft")
    if not (draft and draft.get("state") == "attach"):
        await update.message.reply_text(
            "사진은 메일 작성 중 '첨부할 파일' 단계에서만 첨부할 수 있어요. /mail로 메일 작성을 시작해보세요."
        )
        return

    photo = update.message.photo[-1]  # 가장 고화질
    try:
        tg_file = await photo.get_file()
        file_bytes = bytes(await tg_file.download_as_bytearray())
    except Exception:
        logger.exception("사진 다운로드 중 오류")
        await update.message.reply_text("사진을 받아오는 중 오류가 발생했어요. 다시 보내주세요.")
        return

    n = len(draft.get("attachments", [])) + 1
    filename = f"사진_{n}.jpg"
    draft.setdefault("attachments", []).append((filename, file_bytes))
    await update.message.reply_text(
        f"📎 첨부됨: {filename} (총 {n}개)\n계속 보내시거나, 다 되셨으면 '없음'/'완료'라고 입력해주세요."
    )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return

    doc = update.message.document
    filename = doc.file_name or "attachment"

    # 메일 작성 중 첨부파일 단계면, 파일 종류 상관없이 메일 첨부로 처리
    draft = context.user_data.get("mail_draft")
    if draft and draft.get("state") == "attach":
        try:
            tg_file = await doc.get_file()
            file_bytes = bytes(await tg_file.download_as_bytearray())
        except Exception:
            logger.exception("첨부파일 다운로드 중 오류")
            await update.message.reply_text("파일을 받아오는 중 오류가 발생했어요. 다시 보내주세요.")
            return
        draft.setdefault("attachments", []).append((filename, file_bytes))
        n = len(draft["attachments"])
        await update.message.reply_text(
            f"📎 첨부됨: {filename} (총 {n}개)\n계속 보내시거나, 다 되셨으면 '없음'/'완료'라고 입력해주세요."
        )
        return

    if not filename.lower().endswith(".xlsx"):
        await update.message.reply_text(
            "죄송해요, 지금은 정산양식 엑셀(.xlsx) 파일만 처리할 수 있어요."
        )
        return

    # 서식(PDF)은 이제 브랜드마다 따로 있고 브랜드별로 나중에 확인하니, 여기서는 모든 브랜드가
    # 공통으로 필요한 폰트 파일만 확인함(폰트가 없으면 글자 자체를 못 그리므로 확실한 차단 사유).
    if not os.path.exists(CERT_FONT_PATH):
        await update.message.reply_text("가입증명서 생성 기능이 아직 설정되지 않았어요.")
        return

    await update.message.reply_text("엑셀을 확인하고 있어요...")

    try:
        tg_file = await doc.get_file()
        file_bytes = bytes(await tg_file.download_as_bytearray())
    except Exception:
        logger.exception("엑셀 파일 다운로드 중 오류")
        await update.message.reply_text("파일을 받아오는 중 오류가 발생했어요.")
        return

    result = await _sync_and_notify(context.bot, update.effective_chat.id, file_bytes)
    if result is None:
        await update.message.reply_text(
            "이 엑셀 형식은 알아보지 못했어요. 각 시트 A1 셀에 '*. 브랜드명'이 적힌 정산양식인지 확인해주세요."
        )


def _send_email(to_addr: str, subject: str, body: str, attachments: list | None = None) -> None:
    """attachments: [(파일명, 파일바이트), ...]"""
    if attachments:
        msg = MIMEMultipart()
        msg.attach(MIMEText(body))
        for filename, file_bytes in attachments:
            part = MIMEApplication(file_bytes, Name=filename)
            part["Content-Disposition"] = f'attachment; filename="{filename}"'
            msg.attach(part)
    else:
        msg = MIMEText(body)

    msg["Subject"] = subject
    msg["From"] = MAIL_FROM_ADDRESS
    msg["To"] = to_addr
    if MAIL_REPLY_TO:
        # 보낸사람 계정은 그대로 두고, 상대방이 "답장"을 누르면 유진님 실제 업무 메일로
        # 가도록 함(구글 별칭 인증 없이도 바로 적용됨)
        msg["Reply-To"] = MAIL_REPLY_TO

    # 로그인은 항상 GMAIL_ADDRESS 계정으로 하고, 보낸사람 표시만 MAIL_FROM_ADDRESS로 바뀜
    # (구글의 "다른 이메일 주소로 보내기" 별칭 인증이 되어 있어야 함)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_ADDRESS, [to_addr], msg.as_string())


_gdrive_service = None  # 매번 새로 로그인하지 않도록 한 번 만든 클라이언트를 재사용


def _get_gdrive_service():
    """구글 드라이브 서비스 계정 클라이언트를 만들어서 반환함(설정 안 돼있으면 None).
    유진님이 미리 이 서비스 계정한테 특정 드라이브 폴더를 '공유'해둬야, 그 폴더 안에
    파일을 넣을 수 있음(유진님 로그인/인증 절차 없이 서버에서 바로 동작)."""
    global _gdrive_service
    if _gdrive_service is not None:
        return _gdrive_service
    if not (GDRIVE_SERVICE_ACCOUNT_JSON and GDRIVE_FOLDER_ID and _gdrive_build):
        return None
    try:
        info = json.loads(GDRIVE_SERVICE_ACCOUNT_JSON)
        creds = _gdrive_service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/drive"]
        )
        _gdrive_service = _gdrive_build("drive", "v3", credentials=creds)
        return _gdrive_service
    except Exception:
        logger.exception("구글 드라이브 연결 중 오류")
        return None


def _upload_to_drive(filename: str, file_bytes: bytes) -> bool:
    """가입증명서를 유진님이 지정해둔 구글 드라이브 폴더에 올림. 그 폴더가 유진님 PC의
    구글 드라이브 동기화 폴더 안에 있으면, PC에도 자동으로 똑같이 저장됨. 설정이 안 돼
    있거나 실패하면 조용히 False만 반환(가입증명서 발급 자체는 그대로 진행돼야 하므로)."""
    service = _get_gdrive_service()
    if service is None:
        return False
    try:
        media = _gdrive_media(file_bytes, mimetype="application/pdf")
        service.files().create(
            body={"name": filename, "parents": [GDRIVE_FOLDER_ID]},
            media_body=media,
            fields="id",
        ).execute()
        return True
    except Exception:
        logger.exception("구글 드라이브 업로드 중 오류")
        return False


async def test_mail_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """실제 매장 데이터 없이, 가짜 샘플 매장으로 가입증명서 이메일 발송을 미리 테스트해봄
    (문구/서명/첨부가 실제로 어떻게 보이는지 확인용). 사용법: /testmail <받는이메일>
    (안 적으면 유진님 실제 메일(MAIL_REPLY_TO)로 감)"""
    if not is_allowed(update):
        return
    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        await update.message.reply_text("메일 발송 기능이 설정되어 있지 않아요.")
        return
    to_addr = context.args[0] if context.args else MAIL_REPLY_TO
    if not to_addr:
        await update.message.reply_text("사용법: /testmail <받는이메일>\n예: /testmail yujin.hwang@daonins.co.kr")
        return

    sample = {
        "policy_no": DEFAULT_POLICY_NO,
        "store_code": "TEST0000",
        "store_name": "테스트매장",
        "address": "서울특별시 서초구 강남대로 375 서초현대타워 (테스트용 샘플 주소)",
        "stock_amt": 100000000,
        "facility_amt": 10000000,
        "building_amt": 0,
        "premium": 12345,
        "start_date": "2026.01.01",
        "end_date": "2026.12.31",
        "start_date_yymmdd": "260101",
        "has_property": True,
        "has_liability": True,
    }
    try:
        pdf_bytes = _build_certificate_pdf(sample, "트레몰로")
    except Exception:
        logger.exception("테스트 가입증명서 생성 중 오류")
        await update.message.reply_text("⚠️ 테스트 가입증명서 생성에 실패했어요.")
        return

    out_name = "TEST0000_테스트매장_260101.pdf"
    body = _CERT_EMAIL_BODY_TEMPLATE.format(store_name=sample["store_name"])
    try:
        _send_email(
            to_addr,
            subject="[테스트] 테스트매장 가입증명서",
            body=body,
            attachments=[(out_name, pdf_bytes)],
        )
        await update.message.reply_text(f"✅ 테스트 이메일을 {to_addr}로 보냈어요. 메일함에서 확인해보세요.")
    except Exception:
        logger.exception("테스트 이메일 발송 중 오류")
        await update.message.reply_text("⚠️ 테스트 이메일 발송에 실패했어요.")


async def mail_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        return

    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        await update.message.reply_text("메일 발송 기능이 설정되어 있지 않아요.")
        return

    if not context.args:
        await update.message.reply_text(
            "사용법: /mail <받는사람이메일>\n예: /mail friend@example.com"
        )
        return

    to_addr = context.args[0]
    context.user_data["mail_draft"] = {"to": to_addr, "state": "subject"}
    await update.message.reply_text(
        f"보낸사람: {MAIL_FROM_ADDRESS}\n받는사람: {to_addr}\n메일 제목을 입력해주세요."
    )


def _all_registered_store_rows() -> dict:
    """등록된 모든 브랜드의 '신규매장' 시트를 다 뒤져서, 매장명 -> 그 매장이 있던 행(들) 정보를
    모아둠. '가입증명서를 다시 만들어줘' 같은 요청이 오면 여기서 검색해서 다시 만듦."""
    result: dict[str, list[dict]] = {}
    if not os.path.isdir(MASTERS_DIR):
        return result
    for fname in os.listdir(MASTERS_DIR):
        if not fname.lower().endswith(".xlsx"):
            continue
        brand = fname[:-5]
        path = os.path.join(MASTERS_DIR, fname)
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            wb_values = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        new_sheets = _find_type_sheets(wb, "신규매장")
        new_sheets_values = _find_type_sheets(wb_values, "신규매장")
        for sub_type, ws in new_sheets.items():
            ws_values = new_sheets_values.get(sub_type)
            if ws_values is None:
                continue
            header, min_row = _build_header_map(ws)
            if not header:
                continue
            rate1 = _extract_rate(ws, header, "연간재물보험료", r"\*([\d.]+)%", 0.0665, min_row)
            rate2 = _extract_rate(ws, header, "연간영업배상보험료", r"\*([\d.]+)", 1793, min_row)
            for vals in _extract_data_rows(ws_values, header, min_row, formula_ws=ws, wb_values=wb_values):
                name_raw = str(vals.get("매장명") or "")
                for seg in re.split(r"\n+", name_raw):
                    seg = seg.strip()
                    if not seg:
                        continue
                    result.setdefault(seg, []).append({
                        "brand": brand,
                        "sub_type": sub_type,
                        "vals": vals,
                        "rate1": rate1,
                        "rate2": rate2,
                    })
    return result


_CERT_LOOKUP_RE = re.compile(r"가입\s*증명서")
_CERT_LOOKUP_VERBS = ("찾아", "다시", "재발급", "재전송", "보내줘", "보내주세요", "올려줘", "올려주세요", "보여줘")


async def _handle_cert_lookup_request(update: Update, context: ContextTypes.DEFAULT_TYPE, user_text: str) -> bool:
    """'계산점 가입증명서 찾아줘'처럼, 이미 등록된 매장의 가입증명서를 다시 보내달라는
    자연스러운 문장을 감지해서 처리함. 새 명령어를 안 외우고 그냥 말로 요청할 수 있게 함.
    True를 반환하면 이미 답장까지 다 처리한 것이므로 일반 대화(Claude)로 안 넘어감."""
    if not _CERT_LOOKUP_RE.search(user_text):
        return False
    if not any(v in user_text for v in _CERT_LOOKUP_VERBS):
        return False

    chat_id = update.effective_chat.id
    store_index = _all_registered_store_rows()
    if not store_index:
        return False

    # 문장 안에 등록된 매장명이 포함돼 있는지 확인. 여러 매장명이 겹쳐 걸리면(예: '점'으로
    # 끝나는 짧은 이름이 긴 이름의 일부인 경우) 가장 긴 이름을 우선함
    matches = [name for name in store_index if name and name in user_text]
    if not matches:
        return False
    matches.sort(key=len, reverse=True)
    store_name = matches[0]
    entries = store_index[store_name]

    # 같은 이름으로 여러 건이 있으면(갱신/재등록 등) 접수일자가 가장 최근인 걸 보내줌
    def _recv_key(entry):
        return _parse_date_val(entry["vals"].get("접수일자")) or dt.date.min

    entries.sort(key=_recv_key, reverse=True)
    chosen = entries[0]

    try:
        vals = chosen["vals"]
        cert_vals = _compute_new_store_cert_values(vals, chosen["rate1"], chosen["rate2"])
        address = str(vals.get("매장주소") or "").strip()
        address = re.sub(r"\s*\n\s*", " ", address)
        address = re.sub(r"(?<=\S)\(", " (", address)
        policy_numbers = _load_policy_numbers()
        store = {
            "policy_no": policy_numbers.get(chosen["brand"]) or DEFAULT_POLICY_NO,
            "store_code": str(vals.get("매장코드") or "").strip(),
            "store_name": str(vals.get("매장명") or "").strip(),
            "address": address,
            **cert_vals,
        }
        pdf_bytes = _build_certificate_pdf(store, chosen["brand"])
    except Exception:
        logger.exception("가입증명서 재검색/재생성 중 오류")
        await update.message.reply_text(f"⚠️ '{store_name}' 가입증명서를 다시 만드는 중 오류가 발생했어요.")
        return True

    out_name = f"{store['store_code']}_{store['store_name']}_{store['start_date_yymmdd']}.pdf"
    extra = f" (같은 이름으로 {len(entries)}건 등록되어 있어서 가장 최근 걸로 보내드려요)" if len(entries) > 1 else ""
    await context.bot.send_document(
        chat_id=chat_id,
        document=io.BytesIO(pdf_bytes),
        filename=out_name,
        caption=f"📄 {store['store_name']} 가입증명서{extra}",
    )
    return True


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_allowed(update):
        await update.message.reply_text("죄송해요, 이 봇은 개인용이라 사용할 수 없어요.")
        return

    chat_id = update.effective_chat.id
    user_text = update.message.text

    # 집/회사 위치를 텍스트(주소/장소명)로 저장하려는 중이면 그쪽으로 처리
    saving = context.user_data.get("saving_place")
    if saving:
        place = await _resolve_place(user_text, {})  # '집'/'회사' 문구는 여기선 의미 없으니 빈 places
        if not place:
            await update.message.reply_text("위치를 찾지 못했어요. 다른 이름으로 다시 입력하거나 위치를 공유해주세요.")
            return
        places = _load_places()
        places[saving] = place
        _save_places(places)
        context.user_data.pop("saving_place", None)
        label = "집" if saving == "home" else "회사"
        await update.message.reply_text(f"✅ {label} 위치를 저장했어요: {place.get('name')}")
        return

    # 길찾기(출발지/도착지) 입력 중이면 그쪽으로 처리
    if context.user_data.get("route_query"):
        state = context.user_data["route_query"]["state"]
        place = await _resolve_place(user_text, _load_places())
        if not place:
            if state == "await_current_location":
                await update.message.reply_text(
                    "위치를 찾지 못했어요. 📎(첨부) 버튼으로 정확한 현재 위치를 공유해주시거나, "
                    "계신 곳 이름을 다시 알려주세요."
                )
            else:
                await update.message.reply_text("위치를 찾지 못했어요. 다른 이름으로 다시 입력하거나 위치를 공유해주세요.")
            return
        await _advance_route_query(update, context, place)
        return

    # 메일 작성 중이면(제목/내용/확인 단계) 일반 대화 대신 메일 작성 흐름으로 처리
    draft = context.user_data.get("mail_draft")
    if draft:
        state = draft["state"]

        if state == "subject":
            draft["subject"] = user_text
            draft["state"] = "body"
            await update.message.reply_text("메일 내용을 입력해주세요.")
            return

        if state == "body":
            draft["body"] = user_text
            draft["state"] = "attach"
            draft["attachments"] = []
            await update.message.reply_text(
                "첨부할 파일이 있으면 지금 보내주세요 (여러 개 가능해요, 엑셀/PDF/워드/한글 파일, 사진 등).\n"
                "다 보내셨으면 '없음' 또는 '완료'라고 입력해주세요."
            )
            return

        if state == "attach":
            if user_text.strip() in ("없음", "완료", "no", "done", "skip"):
                draft["state"] = "confirm"
                n = len(draft.get("attachments", []))
                attach_line = f"첨부파일: {n}개\n" if n else ""
                preview = (
                    f"보낸사람: {MAIL_FROM_ADDRESS}\n"
                    f"받는사람: {draft['to']}\n"
                    f"제목: {draft['subject']}\n"
                    f"{attach_line}\n"
                    f"{draft['body']}\n\n"
                    "이대로 보낼까요? '네' 또는 '아니오'로 답해주세요."
                )
                await update.message.reply_text(preview)
            else:
                await update.message.reply_text(
                    "파일을 보내시거나, 다 되셨으면 '없음'/'완료'라고 입력해주세요."
                )
            return

        if state == "confirm":
            context.user_data.pop("mail_draft", None)
            if user_text.strip() in ("네", "예", "ㅇㅇ", "y", "yes", "Y"):
                try:
                    _send_email(
                        draft["to"], draft["subject"], draft["body"],
                        attachments=draft.get("attachments") or None,
                    )
                    await update.message.reply_text("✅ 메일을 보냈어요!")
                except Exception:
                    logger.exception("메일 발송 중 오류")
                    await update.message.reply_text("❌ 메일 발송에 실패했어요. 잠시 후 다시 시도해주세요.")
            else:
                await update.message.reply_text("메일 발송을 취소했어요.")
            return

    # '계산점 가입증명서 찾아줘'처럼 자연스러운 문장으로 재발급을 요청한 경우 감지해서 처리
    if await _handle_cert_lookup_request(update, context, user_text):
        return

    # 자연스러운 문장으로 길찾기를 요청한 경우 감지해서 처리
    if await _handle_natural_route_request(update, context, user_text):
        return

    history = conversations.setdefault(chat_id, [])
    history.append({"role": "user", "content": user_text})
    history = history[-MAX_HISTORY:]

    await context.bot.send_chat_action(chat_id=chat_id, action="typing")

    system_prompt = SYSTEM_PROMPT
    loc = context.user_data.get("location")
    if loc:
        loc_address = loc.get("address") if isinstance(loc, dict) else loc
        system_prompt += (
            f"\n\n참고: 사용자의 최근 공유 위치는 '{loc_address}'입니다. "
            "근처 맛집/장소 등 위치 기반 질문에는 이 정보를 활용해 웹 검색하세요."
        )

    try:
        response = client.messages.create(
            model=MODEL_NAME,
            max_tokens=1024,
            system=system_prompt,
            messages=history,
            tools=CHAT_TOOLS,
        )
        # 도구 사용(검색) 결과가 섞여 있을 수 있어 텍스트 블록만 모아서 답변으로 사용
        reply_text = "".join(
            block.text for block in response.content if block.type == "text"
        ).strip()
        if not reply_text:
            reply_text = "죄송해요, 답변을 생성하지 못했어요. 다시 한번 물어봐주세요."
    except Exception:
        logger.exception("Claude API 호출 중 오류")
        reply_text = "죄송해요, 답변을 만드는 중에 오류가 발생했어요. 잠시 후 다시 시도해주세요."

    history.append({"role": "assistant", "content": reply_text})
    conversations[chat_id] = history

    await update.message.reply_text(reply_text)


def _decode_mime_words(s: str) -> str:
    if not s:
        return ""
    decoded = decode_header(s)
    return "".join(
        (t.decode(enc or "utf-8", errors="ignore") if isinstance(t, bytes) else t)
        for t, enc in decoded
    )


def _get_email_body(msg: email.message.Message) -> str:
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = str(part.get("Content-Disposition"))
            if ctype == "text/plain" and "attachment" not in disp:
                charset = part.get_content_charset() or "utf-8"
                try:
                    return part.get_payload(decode=True).decode(charset, errors="ignore")
                except Exception:
                    return ""
        return ""
    else:
        charset = msg.get_content_charset() or "utf-8"
        try:
            return msg.get_payload(decode=True).decode(charset, errors="ignore")
        except Exception:
            return ""


def _extract_xlsx_attachments(msg: email.message.Message) -> list[tuple[str, bytes]]:
    """메일에서 .xlsx 첨부파일만 골라 (파일명, 파일바이트) 목록으로 반환"""
    attachments = []
    if not msg.is_multipart():
        return attachments
    for part in msg.walk():
        filename = part.get_filename()
        if not filename:
            continue
        filename = _decode_mime_words(filename)
        if not filename.lower().endswith(".xlsx"):
            continue
        try:
            payload = part.get_payload(decode=True)
        except Exception:
            continue
        if payload:
            attachments.append((filename, payload))
    return attachments


async def check_new_mail(context: ContextTypes.DEFAULT_TYPE) -> None:
    global last_uid_seen

    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        return

    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        imap.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        imap.select("INBOX")

        status, data = imap.uid("search", None, "ALL")
        uids = data[0].split()
        if not uids:
            imap.logout()
            return

        latest_uid = int(uids[-1])

        if last_uid_seen is None:
            # 처음 실행될 때는 지금 시점만 기준으로 잡고, 과거 메일은 알리지 않음
            last_uid_seen = latest_uid
            imap.logout()
            return

        new_uids = [uid for uid in uids if int(uid) > last_uid_seen]

        for uid in new_uids:
            status, msg_data = imap.uid("fetch", uid, "(RFC822)")
            if not msg_data or msg_data[0] is None:
                continue
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            subject = _decode_mime_words(msg.get("Subject", "(제목 없음)"))
            sender = _decode_mime_words(msg.get("From", "(발신자 알 수 없음)"))
            body = _get_email_body(msg)[:2000]

            try:
                response = client.messages.create(
                    model=MODEL_NAME,
                    max_tokens=300,
                    system="이메일 내용을 한국어로 3줄 이내로 간결하게 요약해줘. 핵심만 전달해.",
                    messages=[
                        {
                            "role": "user",
                            "content": f"보낸사람: {sender}\n제목: {subject}\n본문:\n{body}",
                        }
                    ],
                )
                summary = response.content[0].text
            except Exception:
                logger.exception("메일 요약 중 오류")
                summary = "(요약 생성 실패)"

            text = f"📬 새 메일 도착\n\n보낸사람: {sender}\n제목: {subject}\n\n요약:\n{summary}"
            await context.bot.send_message(chat_id=ALLOWED_USER_ID, text=text)

            # 정산양식 엑셀 첨부파일이 있으면 자동으로 브랜드 통합파일과 동기화하고,
            # 새로 만들어진 가입증명서는 이 메일을 보낸 사람에게도 자동으로 보내줌
            sender_email = email.utils.parseaddr(sender)[1] or None
            for att_filename, att_bytes in _extract_xlsx_attachments(msg):
                try:
                    result = await _sync_and_notify(
                        context.bot, ALLOWED_USER_ID, att_bytes, requester_email=sender_email
                    )
                except Exception:
                    logger.exception("메일 첨부 엑셀 처리 중 오류")
                    continue
                if result is None:
                    # 정산양식 형식이 아닌 일반 첨부파일일 수 있으니 조용히 넘어감
                    logger.info("메일 첨부파일 '%s'은 정산양식 형식이 아니라 건너뜀", att_filename)

        last_uid_seen = latest_uid
        imap.logout()
    except Exception:
        logger.exception("메일 확인 중 오류")


def main() -> None:
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CommandHandler("remind", remind))
    app.add_handler(CommandHandler("mail", mail_command))
    app.add_handler(CommandHandler("testmail", test_mail_command))
    app.add_handler(CommandHandler("route", route_start))
    app.add_handler(CommandHandler("sethome", set_home))
    app.add_handler(CommandHandler("setwork", set_work))
    app.add_handler(CommandHandler("towork", commute_to_work))
    app.add_handler(CommandHandler("tohome", commute_to_home))
    app.add_handler(CommandHandler("setpolicy", set_policy_command))
    app.add_handler(CommandHandler("brands", list_brands_command))
    app.add_handler(CommandHandler("resetbrand", reset_brand_command))
    app.add_handler(CommandHandler("setbrandname", set_brand_name_command))
    app.add_handler(CommandHandler("sendmaster", send_master_command))
    app.add_handler(CommandHandler("setbrandalias", set_brand_alias_command))
    app.add_handler(CommandHandler("setcerttemplate", set_cert_template_command))
    app.add_handler(CommandHandler("setcontact", set_contact_command))
    app.add_handler(CommandHandler("contacts", list_contacts_command))
    app.add_handler(CommandHandler("pending", list_pending_command))
    app.add_handler(CallbackQueryHandler(handle_cert_confirmation))
    app.add_handler(MessageHandler(filters.LOCATION, handle_location))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    if GMAIL_ADDRESS and GMAIL_APP_PASSWORD:
        app.job_queue.run_repeating(
            check_new_mail, interval=MAIL_CHECK_INTERVAL, first=10
        )
        logger.info("이메일 확인 작업이 등록되었습니다 (%d초 간격).", MAIL_CHECK_INTERVAL)
    else:
        logger.info("GMAIL_ADDRESS/GMAIL_APP_PASSWORD가 없어 이메일 확인 기능은 꺼져 있습니다.")

    logger.info("봇을 시작합니다...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
